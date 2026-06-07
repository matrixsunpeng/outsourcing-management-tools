"""Data processing: read export Excel, filter, deduplicate."""

import re
from datetime import date, datetime
from pathlib import Path

import openpyxl

from config import Config


def read_exported_file(filepath: str) -> list[dict]:
    """Read exported Excel file (.xls or .xlsx) and return list of record dicts.
    Detects format by file header, not extension (IMS exports .xls with .xlsx extension)."""
    # Detect format by header magic bytes
    with open(filepath, 'rb') as f:
        header = f.read(8)

    # OLE2 signature: D0 CF 11 E0 A1 B1 1A E1 → .xls format
    is_ole2 = header[:4] == b'\xd0\xcf\x11\xe0'
    # ZIP signature: PK\x03\x04 → .xlsx format
    is_zip = header[:2] == b'PK'

    if is_ole2:
        return _read_xls(filepath)
    elif is_zip:
        return _read_xlsx(filepath)
    else:
        # Try both, starting with xlrd
        try:
            return _read_xls(filepath)
        except Exception:
            return _read_xlsx(filepath)


def _read_xls(filepath: str) -> list[dict]:
    """Read .xls (OLE2) format using xlrd.
    IMS export format: Row0=title, Row1=headers, data starts at Row2."""
    import xlrd
    wb = xlrd.open_workbook(filepath, formatting_info=False)
    ws = wb.sheet_by_index(0)

    if ws.nrows < 2:
        return []

    # Detect header row: IMS exports have Row0 as title, Row1 as column names
    # Row0 typically has a single merged cell like "人员面试评价表 [2026-05-26]"
    header_row = 0
    non_empty_in_row0 = sum(1 for c in range(ws.ncols) if ws.cell_value(0, c))
    non_empty_in_row1 = sum(1 for c in range(ws.ncols) if ws.cell_value(1, c))

    if non_empty_in_row1 > non_empty_in_row0:
        header_row = 1  # Row 1 is the real header

    headers = [str(ws.cell_value(header_row, c)).strip() for c in range(ws.ncols)]
    records = []

    for r in range(header_row + 1, ws.nrows):
        record = {}
        for c in range(min(ws.ncols, len(headers))):
            key = headers[c]
            if key:
                cell = ws.cell(r, c)
                value = _normalize_xlrd_value(cell, wb)
                record[key] = value
        if record:
            records.append(record)

    return records


def _read_xlsx(filepath: str) -> list[dict]:
    """Read .xlsx format using openpyxl."""
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        wb.close()
        return []

    headers = [str(h).strip() if h else "" for h in rows[0]]
    records = []

    for row in rows[1:]:
        record = {}
        for i, value in enumerate(row):
            if i >= len(headers):
                break
            key = headers[i]
            if key:
                record[key] = _normalize_value(value)
        if record:
            records.append(record)

    wb.close()
    return records


def filter_by_supplier(records: list[dict]) -> tuple[list[dict], list[dict]]:
    """Filter records by allowed suppliers (fuzzy match).
    Returns (matched, rejected).
    """
    allowed = Config.ALLOWED_SUPPLIERS
    matched = []
    rejected = []

    supplier_key = None
    for key in ("供应商/外包商", "供应商", "外包商"):
        if records and records[0].get(key) is not None:
            supplier_key = key
            break

    if not supplier_key:
        # If no supplier column found, return all
        return list(records), []

    for record in records:
        supplier = str(record.get(supplier_key, "")).strip()
        if _fuzzy_match_any(supplier, allowed):
            matched.append(record)
        else:
            rejected.append(record)

    return matched, rejected


def deduplicate(records: list[dict], existing_ids: set[str]) -> tuple[list[dict], int]:
    """Remove records whose 身份证号 already exists in the table.
    Returns (new_records, skipped_count).
    """
    id_keys = ("身份证号", "身份证号码", "身份证")
    new_records = []
    skipped = 0

    for record in records:
        id_no = None
        for key in id_keys:
            id_no = record.get(key, "")
            if id_no:
                break
        id_no = str(id_no).strip()
        if id_no and id_no in existing_ids:
            skipped += 1
        else:
            new_records.append(record)
            if id_no:
                existing_ids.add(id_no)

    return new_records, skipped


def map_to_bitable_fields(records: list[dict]) -> list[dict]:
    """Map Excel column names to Bitable field names."""
    excel_map = Config.EXCEL_TO_BITABLE_MAP
    date_fields = Config.DATE_FIELDS
    number_fields = Config.NUMBER_FIELDS
    result = []

    for record in records:
        mapped = {}
        for excel_col, value in record.items():
            bitable_field = excel_map.get(excel_col)
            if bitable_field is None:
                continue  # Skip unmapped columns
            if value is None or value == "":
                # Skip empty values for optional fields
                continue
            if bitable_field in date_fields:
                value = _to_date_value(value)
            elif bitable_field in number_fields:
                value = _to_number_value(value)
            else:
                value = str(value).strip()
            mapped[bitable_field] = value
        if mapped:
            result.append(mapped)

    return result


def _normalize_xlrd_value(cell, workbook) -> str | float | int:
    """Normalize an xlrd cell value, converting dates properly."""
    import xlrd
    if cell.ctype == xlrd.XL_CELL_DATE:
        # xlrd dates are floats - convert using workbook's datemode
        try:
            dt_tuple = xlrd.xldate_as_tuple(cell.value, workbook.datemode)
            return f"{dt_tuple[0]:04d}-{dt_tuple[1]:02d}-{dt_tuple[2]:02d}"
        except Exception:
            return str(cell.value)
    elif cell.ctype == xlrd.XL_CELL_EMPTY:
        return ""
    elif cell.ctype == xlrd.XL_CELL_NUMBER:
        return cell.value
    return str(cell.value).strip()


def _normalize_value(value):
    """Normalize cell value from openpyxl."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, (int, float)):
        return value
    return str(value).strip()


def _to_date_value(value) -> int | None:
    """Convert date string to Feishu timestamp in milliseconds."""
    if not value:
        return None
    if isinstance(value, (int, float)):
        return int(value)
    s = str(value).strip()
    if not s:
        return None
    try:
        dt = datetime.strptime(s, "%Y-%m-%d")
        return int(dt.timestamp() * 1000)
    except ValueError:
        pass
    try:
        dt = datetime.strptime(s, "%Y-%m-%d %H:%M:%S")
        return int(dt.timestamp() * 1000)
    except ValueError:
        return None


def _to_number_value(value) -> float | None:
    """Convert value to number for Feishu."""
    if not value:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip().replace(",", "").replace("，", "")
    try:
        return float(s)
    except ValueError:
        return None


def _fuzzy_match_any(text: str, patterns: list[str]) -> bool:
    """Check if text contains any of the patterns."""
    for pattern in patterns:
        if pattern in text or text in pattern:
            return True
    return False
