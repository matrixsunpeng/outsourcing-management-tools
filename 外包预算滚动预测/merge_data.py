#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
数据合并工具（安全版 v4）
将下载的计提结算/外包合同数据增补到目标Excel文件中。

安全机制：
- 用openpyxl直接在文件末尾追加行，不重写已有内容（计提结算）
- 外包合同：使用同步合并（按订单月份过滤，范围内用下载数据替换）
- 写入前先备份目标文件
- 文件被占用时报错退出，不损坏原文件
- 只合并downloads目录中的文件（由run_budget_update.py在下载前清空）

主键规则：
- 费用结算单：技术合作订单编号 + 月份 + 单据类型
- 技术合作订单报表：订单编号 + 身份证号

用法：
  python merge_data.py
  python merge_data.py --contract-start 202601 --contract-end 202612
"""
import os
import sys
import glob
import shutil
import re
import argparse
import pandas as pd
import numpy as np
from openpyxl import load_workbook

# ===== 配置（基于本脚本所在目录） =====
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
TARGET_DIR = BASE_DIR
TARGET_SETTLEMENT = os.path.join(TARGET_DIR, '费用结算单计提与结算金额查询.xlsx')
TARGET_CONTRACT = os.path.join(TARGET_DIR, '技术合作订单查询报表.xlsx')
SETTLEMENT_DOWNLOAD_DIR = os.path.join(BASE_DIR, 'settlement', 'downloads')
CONTRACT_DOWNLOAD_DIR = os.path.join(BASE_DIR, 'contract', 'downloads')

# 文件结构：row0=标题, row1=列名, row2+=数据
HEADER_ROWS = 2

# OLE2 (旧版XLS) 的魔术字节
OLE2_MAGIC = b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1'


def is_ole2(filepath: str) -> bool:
    """检测文件是否为OLE2格式（旧版XLS），而非真正的XLSX"""
    try:
        with open(filepath, 'rb') as f:
            return f.read(8) == OLE2_MAGIC
    except Exception:
        return False


def convert_ole2_to_xlsx(filepath: str) -> bool:
    """
    将OLE2格式的文件转换为真正的XLSX格式。
    用pandas读取全部数据，用openpyxl写入标准xlsx，保留所有行（含标题行）。
    返回True表示已转换，False表示无需转换或失败。
    """
    if not is_ole2(filepath):
        return False
    print(f"[WARN] 检测到OLE2格式，正在转换为XLSX: {filepath}")
    try:
        bak = filepath + '.ole2bak'
        shutil.copy2(filepath, bak)
        df = pd.read_excel(bak, header=None)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for r in range(len(df)):
            for c in range(len(df.columns)):
                v = df.iat[r, c]
                if pd.notna(v):
                    ws.cell(row=r + 1, column=c + 1, value=v)
        wb.save(filepath)
        wb.close()
        os.remove(bak)
        print(f"[INFO] 转换完成: {len(df)} 行 ({os.path.getsize(filepath):,} bytes)")
        return True
    except Exception as e:
        print(f"[ERROR] OLE2转XLSX失败: {e}")
        if os.path.exists(bak):
            shutil.copy2(bak, filepath)
        return False


def safe_backup(filepath: str) -> str:
    """备份目标文件"""
    backup = filepath + '.bak'
    if not os.path.exists(backup):
        shutil.copy2(filepath, backup)
        print(f"[INFO] 已备份: {backup}")
    else:
        print(f"[INFO] 备份已存在，跳过: {backup}")
    return backup


def find_downloads(download_dir: str, prefix: str) -> list:
    """找到downloads目录下所有匹配的xlsx文件，按时间排序"""
    pattern = os.path.join(download_dir, f"{prefix}*.xlsx")
    files = glob.glob(pattern)
    return sorted(files, key=os.path.getmtime)


def append_rows_to_excel(target_path: str, new_rows: list, col_names: list):
    """
    用openpyxl在目标文件末尾追加行。
    不修改已有数据，只在最后一行之后写入新记录。
    """
    wb = load_workbook(target_path)
    ws = wb.active
    last_row = ws.max_row
    print(f"[INFO] 目标文件当前共 {last_row} 行（含标题）")

    col_index = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=col).value
        if val is not None:
            col_index[str(val).strip()] = col

    write_row = last_row + 1
    for row_data in new_rows:
        for col_name, col_num in col_index.items():
            val = row_data.get(col_name)
            if pd.notna(val):
                ws.cell(row=write_row, column=col_num, value=val)
        write_row += 1

    wb.save(target_path)
    wb.close()
    print(f"[SUCCESS] 已追加 {len(new_rows)} 行到第 {last_row + 1} ~ {write_row - 1} 行")
    return len(new_rows)


def merge_to_target(target_path: str, download_dir: str, prefix: str,
                    key_columns: list, label: str):
    """
    合并逻辑：
    1. 读取目标主键（只读，不修改）
    2. 遍历下载文件，找出不在目标中的新记录
    3. 用openpyxl在文件末尾追加新行（不碰已有数据）
    """
    print(f"\n{'='*60}")
    print(f"  合并{label}")
    print(f"{'='*60}")

    dl_files = find_downloads(download_dir, prefix)
    if not dl_files:
        print(f"[SKIP] downloads目录中无{label}文件，跳过")
        return 0
    print(f"[INFO] 找到 {len(dl_files)} 个下载文件")

    if not os.path.exists(target_path):
        print(f"[WARN] 目标文件不存在: {target_path}")
        print(f"[INFO] 将从下载文件创建新目标文件")
        src = dl_files[0]
        if is_ole2(src):
            print(f"[INFO] 下载文件为OLE2格式，转换为XLSX")
        df_first_all = pd.read_excel(src, header=None)
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        for r in range(len(df_first_all)):
            for c in range(len(df_first_all.columns)):
                v = df_first_all.iat[r, c]
                if pd.notna(v):
                    ws.cell(row=r + 1, column=c + 1, value=v)
        wb.save(target_path)
        wb.close()
        if len(dl_files) > 1:
            df_first = pd.read_excel(dl_files[0], header=1)
            first_keys = set()
            for _, row in df_first.iterrows():
                key = tuple(str(row.get(col, '')).strip() for col in key_columns)
                first_keys.add(key)
            col_names = list(df_first.columns)
            all_new = []
            for f in dl_files[1:]:
                df_dl = pd.read_excel(f, header=1)
                for _, row in df_dl.iterrows():
                    key = tuple(str(row.get(col, '')).strip() for col in key_columns)
                    if key not in first_keys:
                        all_new.append(row.to_dict())
                        first_keys.add(key)
            if all_new:
                append_rows_to_excel(target_path, all_new, col_names)
        print(f"[SUCCESS] 已创建目标文件: {target_path}")
        return 0

    print(f"[INFO] 目标文件: {target_path}")
    if is_ole2(target_path):
        if not convert_ole2_to_xlsx(target_path):
            print(f"[ERROR] 无法转换目标文件格式，跳过")
            return 0

    try:
        df_target = pd.read_excel(target_path, header=1)
    except Exception as e:
        print(f"[ERROR] 读取目标文件失败: {e}")
        return 0

    target_count = len(df_target)
    col_names = list(df_target.columns)
    print(f"[INFO] 目标文件现有 {target_count} 条记录")

    target_keys = set()
    for _, row in df_target.iterrows():
        key = tuple(str(row.get(col, '')).strip() for col in key_columns)
        target_keys.add(key)

    all_new = []
    total_dl_rows = 0
    for f in dl_files:
        fname = os.path.basename(f)
        print(f"[INFO] 处理: {fname}")
        try:
            df_dl = pd.read_excel(f, header=1)
        except Exception as e:
            print(f"[WARNING] 读取失败: {fname} ({e})，跳过")
            continue
        dl_rows = len(df_dl)
        total_dl_rows += dl_rows
        new_count = 0
        for _, row in df_dl.iterrows():
            key = tuple(str(row.get(col, '')).strip() for col in key_columns)
            if key not in target_keys:
                all_new.append(row.to_dict())
                target_keys.add(key)
                new_count += 1
        print(f"       {dl_rows} 行，新增 {new_count} 条")

    added = len(all_new)
    print(f"[INFO] 汇总: 下载 {total_dl_rows} 行，新增 {added} 条，重复 {total_dl_rows - added} 条")

    if added == 0:
        print(f"[INFO] 无新增记录，不写入")
        return 0

    safe_backup(target_path)

    try:
        append_rows_to_excel(target_path, all_new, col_names)
    except PermissionError:
        print(f"[ERROR] 文件被占用，无法写入: {target_path}")
        print(f"[ERROR] 请关闭Excel后重试！")
        return 0
    except Exception as e:
        print(f"[ERROR] 追加失败: {e}")
        bak = target_path + '.bak'
        if os.path.exists(bak):
            shutil.copy2(bak, target_path)
            print(f"[INFO] 已从备份恢复")
        return 0

    return added


# ============================================================
# 外包合同同步合并（新规则 v4）
# ============================================================

def parse_month_from_order_id(order_id: str) -> int | None:
    """
    从订单编号中解析月份。
    格式示例: AI-OS-202604-0601-亚信科技CMB-1865A-0003
    提取第3段 YYYYMM → 202604
    """
    if not order_id or not isinstance(order_id, str):
        return None
    order_id = order_id.strip()
    parts = order_id.split('-')
    if len(parts) < 3:
        return None
    month_str = parts[2].strip()
    if re.match(r'^\d{6}$', month_str):
        return int(month_str)
    return None


def load_target_df(target_path: str) -> pd.DataFrame | None:
    """读取目标Excel（header=row1，即第2行）"""
    if not os.path.exists(target_path):
        return None
    if is_ole2(target_path):
        if not convert_ole2_to_xlsx(target_path):
            return None
    try:
        return pd.read_excel(target_path, header=1)
    except Exception as e:
        print(f"[ERROR] 读取目标文件失败: {e}")
        return None


def load_download_dfs(download_dir: str, prefix: str) -> list:
    """读取downloads目录下所有匹配的xlsx，返回DataFrame列表"""
    dl_files = find_downloads(download_dir, prefix)
    dfs = []
    for f in dl_files:
        fname = os.path.basename(f)
        print(f"[INFO] 处理: {fname}")
        try:
            df = pd.read_excel(f, header=1)
            dfs.append(df)
            print(f"       {len(df)} 行")
        except Exception as e:
            print(f"[WARNING] 读取失败: {fname} ({e})，跳过")
    return dfs


def build_key_set(df: pd.DataFrame, key_columns: list) -> set:
    """从DataFrame构建主键集合"""
    keys = set()
    for _, row in df.iterrows():
        key = tuple(str(row.get(col, '')).strip() for col in key_columns)
        keys.add(key)
    return keys


def df_to_records(df: pd.DataFrame) -> list:
    """DataFrame → list of dict（保留原始顺序）"""
    return df.to_dict('records')


def rewrite_target_excel(target_path: str, records: list, label: str):
    """
    用 openpyxl 重写整个Excel文件（仅数据部分，保留原表头结构）。
    records 为最终要保留的全部记录（包含原始+新增）。
    """
    if not records:
        print(f"[INFO] 无数据，跳过写入")
        return 0

    if os.path.exists(target_path):
        if is_ole2(target_path):
            convert_ole2_to_xlsx(target_path)
        wb = load_workbook(target_path)
        ws = wb.active
        header_row_idx = 2
        vals_r2 = [ws.cell(row=2, column=c).value for c in range(1, ws.max_column + 1)]
        r2_text_count = sum(1 for v in vals_r2 if v is not None and isinstance(v, str) and len(str(v).strip()) > 0)
        if r2_text_count < 3:
            for trial_row in range(1, 6):
                vals = [ws.cell(row=trial_row, column=c).value for c in range(1, ws.max_column + 1)]
                text_count = sum(1 for v in vals if v is not None and isinstance(v, str) and len(str(v).strip()) > 0)
                if text_count >= 3:
                    header_row_idx = trial_row
                    break
            else:
                header_row_idx = 2
    else:
        print(f"[WARN] 目标文件不存在，将创建新文件: {target_path}")
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        header_row_idx = 2
        if records:
            keys = list(records[0].keys())
            for ci, k in enumerate(keys, 1):
                ws.cell(row=2, column=ci, value=k)

    if not records:
        wb.save(target_path)
        wb.close()
        return 0

    col_index = {}
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row_idx, column=col).value
        if val is not None:
            col_index[str(val).strip()] = col

    if records:
        record_keys = list(records[0].keys())
        for ci, k in enumerate(record_keys, 1):
            if k not in col_index:
                for c in range(1, ws.max_column + 1):
                    if ws.cell(row=header_row_idx, column=c).value is None:
                        col_index[k] = c
                        break
                else:
                    col_index[k] = ws.max_column + (ci - len([v for v in col_index.values() if v <= ws.max_column]))

    last_data_row = ws.max_row
    if last_data_row > header_row_idx:
        ws.delete_rows(header_row_idx + 1, last_data_row - header_row_idx)

    write_row = header_row_idx + 1
    for record in records:
        for col_name, col_num in col_index.items():
            val = record.get(col_name)
            if pd.notna(val):
                ws.cell(row=write_row, column=col_num, value=val)
        write_row += 1

    added = write_row - (header_row_idx + 1)
    wb.save(target_path)
    wb.close()
    print(f"[SUCCESS] 已写入 {added} 行到: {os.path.basename(target_path)}")
    return added


def merge_contract_sync(target_path: str, download_dir: str, prefix: str,
                        key_columns: list, month_range: tuple | None,
                        label: str) -> int:
    """
    外包合同同步合并（新规则）：

    1. 解析 month_range（起止月份，YYYYMM整数）
    2. 读取原合同清单（含全部历史记录）
    3. 读取下载文件，合并去重
    4. 分类原记录：月份不在范围内保留，在范围内用下载数据替换
    5. 用 openpyxl 重写目标文件

    month_range: (start_month_int, end_month_int)，如 (202601, 202612)
                  None 时退化为纯追加逻辑
    """
    print(f"\n{'='*60}")
    print(f"  同步合并 {label}（按月份过滤 + 同步）")
    print(f"{'='*60}")

    dl_files = find_downloads(download_dir, prefix)
    if not dl_files:
        print(f"[SKIP] downloads目录中无{label}文件，跳过")
        return 0
    print(f"[INFO] 找到 {len(dl_files)} 个下载文件")

    dl_dfs = load_download_dfs(download_dir, prefix)
    if not dl_dfs:
        print(f"[ERROR] 无可用的下载数据，跳过")
        return 0

    df_dl_all = pd.concat(dl_dfs, ignore_index=True)
    df_dl_all.drop_duplicates(subset=key_columns, keep='first', inplace=True)
    df_dl_all.reset_index(drop=True, inplace=True)
    print(f"[INFO] 下载文件合计 {len(df_dl_all)} 条（去重后）")

    dl_key_set = build_key_set(df_dl_all, key_columns)
    dl_records = df_to_records(df_dl_all)

    df_orig = load_target_df(target_path)
    if df_orig is None:
        print(f"[WARN] 目标文件不存在，将用下载数据创建新文件")
        return rewrite_target_excel(target_path, dl_records, label)

    print(f"[INFO] 原合同清单共 {len(df_orig)} 条记录")
    df_orig.reset_index(drop=True, inplace=True)
    orig_records = df_to_records(df_orig)
    orig_keys = build_key_set(df_orig, key_columns)

    if month_range is None:
        print(f"[INFO] 无月份范围参数，退化为追加模式")
        new_records = [r for r in dl_records
                       if tuple(str(r.get(col, '')).strip() for col in key_columns) not in orig_keys]
        if not new_records:
            print(f"[INFO] 无新增记录，不写入")
            return 0
        safe_backup(target_path)
        return append_rows_to_excel(target_path, new_records, list(df_orig.columns))

    start_month, end_month = month_range
    print(f"[INFO] 同步范围: {start_month} ~ {end_month}")

    outside_range = []
    inside_range = []
    outside_keys = set()
    inside_keys = set()

    for record in orig_records:
        order_id = str(record.get('订单编号', '')).strip()
        month = parse_month_from_order_id(order_id)
        key = tuple(str(record.get(col, '')).strip() for col in key_columns)
        if month is None:
            inside_range.append(record)
            inside_keys.add(key)
        elif month < start_month or month > end_month:
            outside_range.append(record)
            outside_keys.add(key)
        else:
            inside_range.append(record)
            inside_keys.add(key)

    print(f"  原记录分类: 范围内 {len(inside_range)} 条，范围外 {len(outside_range)} 条")

    to_delete_keys = inside_keys - dl_key_set
    deleted_count = len(to_delete_keys)
    print(f"  范围内待删除: {deleted_count} 条（原清单有但下载无）")

    to_add_records = [r for r in dl_records
                      if tuple(str(r.get(col, '')).strip() for col in key_columns) not in orig_keys]
    added_count = len(to_add_records)
    print(f"  范围内待新增: {added_count} 条（原清单无但下载有）")

    dl_key_to_record = {}
    for r in dl_records:
        k = tuple(str(r.get(col, '')).strip() for col in key_columns)
        dl_key_to_record[k] = r

    final_records = list(outside_range)

    replaced_count = 0
    for record in inside_range:
        key = tuple(str(record.get(col, '')).strip() for col in key_columns)
        if key in dl_key_set:
            final_records.append(dl_key_to_record[key].copy())
            replaced_count += 1

    final_records.extend(to_add_records)

    if replaced_count:
        print(f"  范围内主键一致已整体替换: {replaced_count} 条")
    print(f"  最终记录数: {len(final_records)} 条（范围外{len(outside_range)} + 范围内一致{replaced_count} + 新增{added_count}）")

    safe_backup(target_path)
    try:
        written = rewrite_target_excel(target_path, final_records, label)
    except PermissionError:
        print(f"[ERROR] 文件被占用，无法写入: {target_path}")
        print(f"[ERROR] 请关闭Excel后重试！")
        bak = target_path + '.bak'
        if os.path.exists(bak):
            shutil.copy2(bak, target_path)
            print(f"[INFO] 已从备份恢复")
        return 0
    except Exception as e:
        print(f"[ERROR] 写入失败: {e}")
        bak = target_path + '.bak'
        if os.path.exists(bak):
            shutil.copy2(bak, target_path)
            print(f"[INFO] 已从备份恢复")
        return 0

    return added_count + replaced_count


def main():
    parser = argparse.ArgumentParser(description='数据合并工具')
    parser.add_argument('--contract-start', '--cs', default='',
                        help='外包合同下载起始月份（YYYYMM格式）')
    parser.add_argument('--contract-end', '--ce', default='',
                        help='外包合同下载结束月份（YYYYMM格式）')
    args = parser.parse_args()

    cs = args.contract_start.strip()
    ce = args.contract_end.strip()
    contract_month_range = None
    if cs and ce:
        try:
            start_month = int(cs)
            end_month = int(ce)
            if start_month <= end_month:
                contract_month_range = (start_month, end_month)
                print(f"[INFO] 外包合同同步月份范围: {start_month} ~ {end_month}")
            else:
                print(f"[WARN] 起始月份 {start_month} > 结束月份 {end_month}，将不执行范围过滤")
        except ValueError:
            print(f"[WARN] 无效的月份格式（需为YYYYMM），将跳过范围过滤")

    total = 0

    total += merge_to_target(
        target_path=TARGET_SETTLEMENT,
        download_dir=SETTLEMENT_DOWNLOAD_DIR,
        prefix='计提结算_',
        key_columns=['技术合作订单编号', '月份', '单据类型'],
        label='计提结算'
    )

    total += merge_contract_sync(
        target_path=TARGET_CONTRACT,
        download_dir=CONTRACT_DOWNLOAD_DIR,
        prefix='外包合同_',
        key_columns=['订单编号', '身份证号'],
        month_range=contract_month_range,
        label='外包合同'
    )

    print(f"\n{'='*60}")
    print(f"  合并完成，共新增/同步 {total} 条记录")
    print(f"{'='*60}")
    return total


if __name__ == '__main__':
    if sys.platform == 'win32':
        os.system('chcp 65001 >nul 2>&1')
        sys.stdout.reconfigure(encoding='utf-8')
    main()
