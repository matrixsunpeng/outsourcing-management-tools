#!/usr/bin/env python3
"""Generate Budget Dashboard HTML from Excel data sources."""
import pandas as pd, numpy as np, json, math, datetime, re, sys, os
from collections import defaultdict

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
OUT_FILE = f'{BASE_DIR}/dashboard.html'

BU_MAP = {
    '亚信科技CMB':'CMB','亚信科技CSC':'CSC','亚信科技CMD':'CMD',
    '亚信科技CUC':'CUC','亚信科技CTC':'CTC','亚信科技ESU':'ESU',
    '亚信科技AID':'AID','亚信科技RIC':'RIC','亚信科技AIO':'AIO','亚信科技CSU':'CSU'
}
BUS = sorted(BU_MAP.values())
ME = [None]+[datetime.date(2026,m,[31,28,31,30,31,30,31,31,30,31,30,31][m-1]) for m in range(1,13)]
PC = [None]+['一月标准单价','二月标准单价','三月标准单价','四月标准单价','五月标准单价','六月标准单价',
            '七月标准单价','八月标准单价','九月标准单价','十月标准单价','十一月标准单价','十二月标准单价']
# 锚点默认值：优先使用命令行参数，否则取本地电脑当前月份首日
if len(sys.argv) > 1:
    try:
        _arg = sys.argv[1].replace('-', '')
        ANCHORAGE = datetime.datetime.strptime(_arg, '%Y%m%d').date()
    except Exception:
        print(f'锚点时间参数格式错误: {sys.argv[1]}，使用默认值')
        ANCHORAGE = datetime.date.today().replace(day=1)
else:
    ANCHORAGE = datetime.date.today().replace(day=1)
ANCHORAGE_S = ANCHORAGE.strftime('%Y%m%d')

def cv(v):
    if v is None or (isinstance(v,float) and (math.isnan(v) if not math.isinf(v) else True)): return 0
    return math.ceil(v)

def cbu(name):
    if pd.isna(name): return None
    s = re.sub(r'[\r\n]','',str(name)).strip()
    s = re.sub(r'[（(][^）)]*[）)]','',s).strip()
    return s if s in BUS else None

print("Reading data sources...")
df_budget = pd.read_excel(f'{BASE_DIR}/26年全面预算.xlsx', header=0)
budget = {}
for _,r in df_budget.iterrows():
    bu=cbu(r['BU'])
    if bu: budget[bu]={'ac':int(r['年度费用（千元）']),'ah':int(r['年化HC（人）'])}

df_fee = pd.read_excel(f'{BASE_DIR}/费用结算单计提与结算金额查询.xlsx', header=1)
df_fee['BU']=df_fee['SBU'].map(BU_MAP); df_fee=df_fee[df_fee['BU'].notna()].copy()

df_order = pd.read_excel(f'{BASE_DIR}/技术合作订单查询报表.xlsx', header=1)
df_order['BU']=df_order['申请部门'].map(BU_MAP); df_order=df_order[df_order['BU'].notna()].copy()
df_order['工作开始时间']=pd.to_datetime(df_order['工作开始时间'],errors='coerce')
df_order['工作结束时间']=pd.to_datetime(df_order['工作结束时间'],errors='coerce')
df_act=df_order[df_order['单据状态']=='审批流程结束'].copy()
oid_pool=dict(zip(df_order['订单编号'].fillna(''),df_order['资源池代码'].fillna('')))

df_inv=pd.read_excel(f'{BASE_DIR}/26年继续投入.xlsx',header=None)
ser=df_inv.iloc[0,2:].tolist()
inv_m=[]
for s in ser:
    if pd.notna(s):
        d=datetime.date(1899,12,30)+datetime.timedelta(days=int(s)); inv_m.append(d.month)

df_pool=pd.read_csv(f'{BASE_DIR}/资源池列表.xls',encoding='utf-8-sig')
df_pool['BU']=df_pool['BU'].map(BU_MAP); df_pool=df_pool[df_pool['BU'].notna()].copy()
df_pool['年度费用预算(元)']=df_pool['年度费用预算(元)'].astype(str).str.replace(',','').astype(float)

print("Processing actual costs by BU...")
ac_bu={bu:{m:0 for m in range(1,13)} for bu in BUS}
for _,r in df_fee.iterrows():
    m=int(r['月份'])%100; v=r['当月计提']
    if pd.notna(v): ac_bu[r['BU']][m]+=v
for bu in BUS:
    for m in range(1,13): ac_bu[bu][m]=cv(ac_bu[bu][m]/1000)

print("Processing actual costs by pool (from 计提结算报表)...")
# 虚拟项目代码 → 资源池代码 映射
vpc_to_pool={}
for _,r in df_pool.iterrows():
    vpc=str(r['虚拟项目代码']).strip() if pd.notna(r.get('虚拟项目代码')) else ''
    pc=str(r['资源池代码']).strip() if pd.notna(r.get('资源池代码')) else ''
    if vpc and pc: vpc_to_pool[vpc]=pc
ac_pl_raw=defaultdict(lambda:defaultdict(float))
for _,r in df_fee.iterrows():
    proj_code=str(r['项目代码']).strip() if pd.notna(r.get('项目代码')) else ''
    pc=vpc_to_pool.get(proj_code)
    if not pc: continue
    m=int(r['月份'])%100; v=r['当月计提']
    if pd.notna(v): ac_pl_raw[pc][m]+=v
ac_pl={}
for pc in ac_pl_raw:
    ac_pl[pc]={}
    for m in range(1,13): ac_pl[pc][m]=cv(ac_pl_raw[pc][m]/1000)
print(f"  {len(ac_pl)} pools with actual cost data")

print("Processing actual HC by BU...")
ah_bu={bu:{m:0 for m in range(1,13)} for bu in BUS}
for m in range(1,13):
    me=pd.Timestamp(ME[m])
    mask=df_act['工作开始时间'].notna()&df_act['工作结束时间'].notna()&(df_act['工作开始时间']<=me)&(df_act['工作结束时间']>=me)
    for bu,c in df_act[mask].groupby('BU').size().items(): ah_bu[bu][m]=int(c)

print("Processing actual HC by pool...")
ah_pl_raw=defaultdict(lambda:defaultdict(int))
for m in range(1,13):
    me=pd.Timestamp(ME[m])
    mask=df_act['工作开始时间'].notna()&df_act['工作结束时间'].notna()&(df_act['工作开始时间']<=me)&(df_act['工作结束时间']>=me)
    for pc,c in df_act[mask].groupby('资源池代码').size().items():
        if pd.notna(pc): ah_pl_raw[pc][m]=int(c)
ah_pl={pc:dict(ah_pl_raw[pc]) for pc in ah_pl_raw}

# ============================================================
# 续签数据：传人员明细到前端，由前端按精确日期实时计算
# ============================================================
print("Building renewal staff list (all statuses)...")
# 续签预测口径：全部记录（不限单据状态），只要在锚点当天在岗即计入
staff_list = []
for _, r in df_order.iterrows():
    ws = r['工作开始时间']
    we = r['工作结束时间']
    if pd.isna(ws) or pd.isna(we):
        continue
    bu = r['BU'] if pd.notna(r.get('BU')) else cbu(r.get('申请部门', ''))
    if not bu:
        continue
    staff_list.append({
        'bu': bu,
        'pool': str(r['资源池代码']) if pd.notna(r.get('资源池代码')) else '',
        'ws': ws.strftime('%Y-%m-%d'),
        'we': we.strftime('%Y-%m-%d'),
        'rate': float(r['费率']) if pd.notna(r.get('费率')) else 0,
        'mp': [float(r[PC[m]]) if pd.notna(r.get(PC[m])) else 0 for m in range(1, 13)]
    })
print(f"  {len(staff_list)} staff records (all statuses)")

# 实际发生人数按池计算仍使用审批流程结束的记录
print("Building actual HC by pool (using 审批流程结束)...")
ah_pl_raw=defaultdict(lambda:defaultdict(int))
for m in range(1,13):
    me=pd.Timestamp(ME[m])
    mask=df_act['工作开始时间'].notna()&df_act['工作结束时间'].notna()&(df_act['工作开始时间']<=me)&(df_act['工作结束时间']>=me)
    for pc,c in df_act[mask].groupby('资源池代码').size().items():
        if pd.notna(pc): ah_pl_raw[pc][m]=int(c)
ah_pl={pc:dict(ah_pl_raw[pc]) for pc in ah_pl_raw}

print("Processing continue invest...")
ch_bu={bu:{m:{'intern':0,'cont':0} for m in range(1,13)} for bu in BUS}
cc_bu={bu:{m:{'intern':0,'cont':0} for m in range(1,13)} for bu in BUS}
for _,r in df_inv.iterrows():
    if r[0]=='BU' or pd.isna(r[0]): continue
    bu=cbu(r[0])
    if not bu: continue
    t=re.sub(r'[\r\n]','',str(r[1])).strip()
    is_int='实习生' in t; is_cost='千元' in t
    k='intern' if is_int else 'cont'
    for i,m in enumerate(inv_m):
        if m is None: continue
        v=r[2+i]
        if pd.notna(v):
            if is_cost: cc_bu[bu][m][k]=cv(v)
            else: ch_bu[bu][m][k]=int(v)

print("Processing pool list (filtering to pools with data)...")
active_pool_codes = df_act['资源池代码'].dropna().unique()
pl_by_bu=defaultdict(list)
for _,r in df_pool.iterrows():
    bu=r['BU']
    if pd.isna(bu): continue
    pc=r['资源池代码']
    if pc not in active_pool_codes: continue
    pl_by_bu[bu].append({'code':pc,'name':r['资源池名称'],
        'org_budget':cv(r['年度费用预算(元)']/1000),
        'ahb':round(float(r['年化人数预算']),1) if pd.notna(r.get('年化人数预算')) else 0})

print("Building JSON...")
data={'buList':BUS,'budget':budget,'ac':ac_bu,'ah':ah_bu,
      'ch':ch_bu,'cc':cc_bu,'pl':{bu:pl_by_bu.get(bu,[]) for bu in BUS},
      'acPl':ac_pl,'ahPl':ah_pl,'staff':staff_list}
js=json.dumps(data,ensure_ascii=False)

print("Generating HTML...")
HTML_TMPL = r'''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>预算管控仪表盘 2026</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+SC:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<style>
*{margin:0;padding:0;box-sizing:border-box}
:root{--bg:#0F172A;--card:rgba(30,41,59,0.7);--border:rgba(59,130,246,0.15);--accent1:#3B82F6;--accent2:#8B5CF6;
  --text:#F8FAFC;--text2:#94A3B8;--text3:#64748B;--green:#22C55E;--yellow:#EAB308;--red:#EF4444;--glow:rgba(59,130,246,0.3)}
body{font-family:'Noto Sans SC',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;overflow-x:hidden}
.bg-deco{position:fixed;top:0;left:0;width:100%;height:100%;pointer-events:none;z-index:0;
  background:radial-gradient(ellipse at 20% 20%,rgba(59,130,246,0.08) 0%,transparent 50%),
  radial-gradient(ellipse at 80% 80%,rgba(139,92,246,0.06) 0%,transparent 50%)}
.header{position:sticky;top:0;z-index:50;backdrop-filter:blur(20px);background:rgba(15,23,42,0.85);
  border-bottom:1px solid var(--border);padding:16px 32px;display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px}
.logo{display:flex;align-items:center;gap:12px}
.logo-icon{width:36px;height:36px;border-radius:10px;background:linear-gradient(135deg,var(--accent1),var(--accent2));
  display:flex;align-items:center;justify-content:center;font-size:18px;box-shadow:0 0 20px var(--glow)}
.logo h1{font-size:22px;font-weight:700;background:linear-gradient(135deg,var(--accent1),var(--accent2));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent}
.logo .badge{background:rgba(59,130,246,0.15);color:var(--accent1);padding:2px 10px;border-radius:12px;font-size:12px;font-weight:600}
.controls{display:flex;align-items:center;gap:12px;flex-wrap:wrap}
.controls label{font-size:13px;color:var(--text2);font-weight:500}
.controls select,.controls input[type=text]{background:rgba(30,41,59,0.9);border:1px solid var(--border);
  color:var(--text);padding:8px 14px;border-radius:8px;font-size:14px;font-family:inherit;outline:none;transition:border .2s}
.controls select:focus,.controls input:focus{border-color:var(--accent1);box-shadow:0 0 0 3px rgba(59,130,246,0.15)}
.controls select{min-width:160px;cursor:pointer}
.controls input{width:110px}
.btn{background:linear-gradient(135deg,var(--accent1),var(--accent2));color:#fff;border:none;padding:8px 20px;
  border-radius:8px;font-size:14px;font-weight:600;cursor:pointer;transition:transform .15s,box-shadow .15s;font-family:inherit}
.btn:hover{transform:translateY(-1px);box-shadow:0 4px 20px var(--glow)}
.btn:active{transform:translateY(0)}
.main{position:relative;z-index:1;padding:24px 32px;max-width:1600px;margin:0 auto}
.section-title{font-size:18px;font-weight:600;margin:28px 0 16px;color:var(--text);display:flex;align-items:center;gap:8px}
.section-title::before{content:'';width:4px;height:20px;border-radius:2px;background:linear-gradient(var(--accent1),var(--accent2))}
.cards{display:grid;grid-template-columns:repeat(auto-fill,minmax(300px,1fr));gap:16px;margin-bottom:8px}
.card{background:var(--card);backdrop-filter:blur(10px);border:1px solid var(--border);border-radius:14px;
  padding:20px;transition:transform .2s,box-shadow .2s;position:relative;overflow:hidden}
.card:hover{transform:translateY(-2px);box-shadow:0 8px 30px rgba(0,0,0,0.3);border-color:rgba(59,130,246,0.3)}
.card .bu-name{font-size:20px;font-weight:700;margin-bottom:12px;background:linear-gradient(135deg,var(--accent1),var(--accent2));
  -webkit-background-clip:text;-webkit-text-fill-color:transparent}
.card .row{display:flex;justify-content:space-between;align-items:center;padding:4px 0;font-size:13px}
.card .row .lbl{color:var(--text2)}
.card .row .val{font-weight:600;font-size:15px}
.card .row .val.predicted{color:var(--accent2)}
.progress-wrap{margin-top:10px}
.progress-label{display:flex;justify-content:space-between;font-size:12px;color:var(--text2);margin-bottom:4px}
.progress-bar{height:8px;background:rgba(255,255,255,0.06);border-radius:4px;overflow:hidden}
.progress-fill{height:100%;border-radius:4px;transition:width .6s ease;min-width:0}
.progress-fill.green{background:linear-gradient(90deg,#22C55E,#4ADE80)}
.progress-fill.yellow{background:linear-gradient(90deg,#EAB308,#FACC15)}
.progress-fill.red{background:linear-gradient(90deg,#EF4444,#F87171)}
.tabs{display:flex;gap:0;margin:20px 0 0;border-bottom:2px solid rgba(255,255,255,0.06)}
.tab{padding:10px 24px;font-size:14px;font-weight:500;color:var(--text3);cursor:pointer;position:relative;transition:color .2s}
.tab:hover{color:var(--text2)}
.tab.active{color:var(--accent1)}
.tab.active::after{content:'';position:absolute;bottom:-2px;left:0;right:0;height:2px;background:linear-gradient(var(--accent1),var(--accent2));border-radius:1px}
.tab-content{display:none;padding:16px 0}
.tab-content.active{display:block}
.tbl-wrap{overflow-x:auto;border-radius:10px;border:1px solid var(--border)}
table{width:100%;border-collapse:collapse;font-size:13px}
thead th{background:rgba(30,41,59,0.95);padding:10px 12px;text-align:center;font-weight:600;color:var(--text2);
  border-bottom:1px solid var(--border);position:sticky;top:0;white-space:nowrap}
thead th:first-child{text-align:left;min-width:80px}
tbody td{padding:9px 12px;text-align:center;border-bottom:1px solid rgba(255,255,255,0.03);white-space:nowrap}
tbody td:first-child{text-align:left;color:var(--text2);font-weight:500}
tbody tr:hover{background:rgba(59,130,246,0.04)}
tbody tr.dim td{color:var(--text3);opacity:.5}
.pool-section{margin-top:8px}
.pool-table thead th{font-size:12px;padding:8px 10px}
.pool-table tbody td{font-size:12px;padding:7px 10px}
.pool-row{cursor:pointer;transition:background .15s}
.pool-row:hover{background:rgba(59,130,246,0.06)!important}
.pool-detail{display:none;background:rgba(15,23,42,0.5)}
.pool-detail.open{display:table-row}
.pool-detail td{padding:0!important}
.pool-detail-inner{padding:12px 16px}
.pool-detail-inner .sub-table{margin-top:8px;font-size:12px}
.pool-detail-inner .sub-table th,.pool-detail-inner .sub-table td{padding:6px 8px;font-size:11px}
.num{font-variant-numeric:tabular-nums}
.tag{display:inline-block;padding:2px 8px;border-radius:4px;font-size:11px;font-weight:600}
.tag-green{background:rgba(34,197,94,0.15);color:#22C55E}
.tag-yellow{background:rgba(234,179,8,0.15);color:#EAB308}
.tag-red{background:rgba(239,68,68,0.15);color:#EF4444}
.empty{text-align:center;color:var(--text3);padding:40px;font-size:14px}
@media(max-width:768px){
  .header{padding:12px 16px}
  .main{padding:16px}
  .cards{grid-template-columns:1fr}
  .controls{width:100%}
}
.fade-in{animation:fadeIn .4s ease}
@keyframes fadeIn{from{opacity:0;transform:translateY(8px)}to{opacity:1;transform:translateY(0)}}
</style>
</head>
<body>
<div class="bg-deco"></div>
<div class="header">
  <div class="logo">
    <div class="logo-icon">&#9776;</div>
    <h1>预算管控仪表盘</h1>
    <span class="badge">2026</span>
  </div>
  <div class="controls">
    <label>BU</label>
    <select id="selBU"><option value="">全部BU合计</option></select>
    <label>锚点时间</label>
    <input type="text" id="inpAnchor" value="__ANCHORAGE__" placeholder="YYYYMMDD" maxlength="8">
    <button class="btn" onclick="render()">查询</button>
    <button class="btn" onclick="exportExcel()" style="background:linear-gradient(135deg,#22C55E,#16A34A)">导出Excel</button>
  </div>
</div>
<div class="main">
  <div class="section-title">BU分析</div>
  <div class="cards" id="cards"></div>
  <div class="section-title">明细数据</div>
  <div class="tabs">
    <div class="tab active" data-tab="actual" onclick="switchTab('actual')">实际发生</div>
    <div class="tab" data-tab="renewal" onclick="switchTab('renewal')">预测续签</div>
    <div class="tab" data-tab="invest" onclick="switchTab('invest')">继续投入</div>
  </div>
  <div class="tab-content active" id="tab-actual"></div>
  <div class="tab-content" id="tab-renewal"></div>
  <div class="tab-content" id="tab-invest"></div>
  <div id="poolSection" class="pool-section" style="display:none">
    <div class="section-title">资源池明细</div>
    <div id="poolContent"></div>
  </div>
</div>
<script>
const D=__JSON_PLACEHOLDER__;
const MNAME=['1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月'];
const sel=document.getElementById('selBU');
D.buList.forEach(function(b){var o=document.createElement('option');o.value=b;o.textContent=b;sel.appendChild(o)});

function gv(obj,m){return (obj&&obj[m])||0}
function fmt(n){return n.toLocaleString('zh-CN')}
function pcolor(pct){if(pct<70)return'green';if(pct<90)return'yellow';return'red'}
function pclass(pct){return'progress-fill '+pcolor(pct)}

// 根据锚点当天在岗人数，汇总续签数据
// 人数：锚点当天在岗（ws<=anchDate && we>=anchDate）
// 费用：锚点当天在岗人员，每月核算费用
function calcRenewal(anchDate){
  var anchM=parseInt(anchDate.substr(5,2));
  var byBU={}, byPool={};
  var hc=0;
  D.staff.forEach(function(s){
    if(s.ws<=anchDate && s.we>=anchDate){
      hc++;
      if(!byBU[s.bu]) byBU[s.bu]={hc:0,cost:{}};
      byBU[s.bu].hc++;
      if(s.pool){
        if(!byPool[s.pool]) byPool[s.pool]={hc:0,cost:{}};
        byPool[s.pool].hc++;
      }
    }
  });
  // 费用：锚点月至12月，按锚点当天在岗的同一批人逐月核算
  for(var mm=anchM;mm<=12;mm++){
    var mi=mm-1;
    D.staff.forEach(function(s){
      if(s.ws<=anchDate && s.we>=anchDate){
        var price=s.mp[mi];
        var cost=Math.ceil(price*s.rate/100/1.06)/1000;
        if(byBU[s.bu]){
          byBU[s.bu].cost[mm]=(byBU[s.bu].cost[mm]||0)+cost;
        }
        if(s.pool && byPool[s.pool]){
          byPool[s.pool].cost[mm]=(byPool[s.pool].cost[mm]||0)+cost;
        }
      }
    });
  }
  Object.keys(byBU).forEach(function(bu){
    Object.keys(byBU[bu].cost).forEach(function(m){ byBU[bu].cost[m]=Math.ceil(byBU[bu].cost[m]); });
  });
  Object.keys(byPool).forEach(function(pc){
    Object.keys(byPool[pc].cost).forEach(function(m){ byPool[pc].cost[m]=Math.ceil(byPool[pc].cost[m]); });
  });
  return {hc:hc, byBU:byBU, byPool:byPool};
}

function calcBU(bu,anchMonth,anchDate,rwData){
  var b=D.budget[bu], ac=D.ac[bu]||{}, ah=D.ah[bu]||{};
  var rw=rwData[bu]||{hc:0,cost:{}};
  var ch=D.ch[bu]||{}, cc=D.cc[bu]||{};
  var totalCost=0, mHC=[];
  for(var m=1;m<=12;m++){
    var cost,hc;
    if(m<anchMonth){
      cost=gv(ac,m)+gv(gv(cc,m),'intern')+gv(gv(cc,m),'cont');
      hc=gv(ah,m)+gv(gv(ch,m),'intern')+gv(gv(ch,m),'cont');
    } else {
      cost=gv(rw.cost,m)+gv(gv(cc,m),'intern')+gv(gv(cc,m),'cont');
      hc=rw.hc+gv(gv(ch,m),'intern')+gv(gv(ch,m),'cont');
    }
    totalCost+=cost; mHC.push(hc);
  }
  var avgHC=Math.ceil(mHC.reduce(function(a,b){return a+b},0)/12);
  var cp=b?(totalCost/b.ac*100).toFixed(1):null;
  var hp=b?(avgHC/b.ah*100).toFixed(1):null;
  return {totalCost:totalCost,avgHC:avgHC,cp:cp,hp:hp,mHC:mHC,ac:ac,ah:ah,rw:rw,ch:ch,cc:cc};
}

function calcAll(anchMonth,anchDate,rwData){
  var tc=0, mHC=new Array(12).fill(0), tac={}, tah={};
  D.buList.forEach(function(bu){
    var r=calcBU(bu,anchMonth,anchDate,rwData); tc+=r.totalCost;
    for(var m=1;m<=12;m++){
      mHC[m-1]+=r.mHC[m-1];
      tac[m]=(tac[m]||0)+gv(r.ac,m)+gv(gv(r.cc,m),'intern')+gv(gv(r.cc,m),'cont');
      tah[m]=(tah[m]||0)+gv(r.ah,m)+gv(gv(r.ch,m),'intern')+gv(gv(r.ch,m),'cont');
    }
  });
  var sumAC=0,sumAH=0;
  D.buList.forEach(function(bu){sumAC+=(D.budget[bu]||{}).ac||0;sumAH+=(D.budget[bu]||{}).ah||0});
  var avgHC=Math.ceil(mHC.reduce(function(a,b){return a+b},0)/12);
  var cp=sumAC?(tc/sumAC*100).toFixed(1):null;
  var hp=sumAH?(avgHC/sumAH*100).toFixed(1):null;
  var rwHC=0;D.buList.forEach(function(bu){rwHC+=(rwData[bu]||{}).hc||0});
  return {totalCost:tc,avgHC:avgHC,cp:cp,hp:hp,mHC:mHC,tac:tac,tah:tah,rwHC:rwHC};
}

function renderCards(bu,anchMonth,anchDate,rwData){
  var el=document.getElementById('cards');
  if(bu){
    var r=calcBU(bu,anchMonth,anchDate,rwData); var b=D.budget[bu];
    el.innerHTML=cardHTML(bu,b,r);
  } else {
    var r=calcAll(anchMonth,anchDate,rwData);
    var sumAC=0,sumAH=0;
    D.buList.forEach(function(b){sumAC+=(D.budget[b]||{}).ac||0;sumAH+=(D.budget[b]||{}).ah||0});
    var h='<div class="cards" style="margin-bottom:16px">';
    h+=cardHTML('合计',{ac:sumAC,ah:sumAH},r);
    h+='</div><div class="cards">';
    D.buList.forEach(function(b){
      var br=calcBU(b,anchMonth,anchDate,rwData); h+=cardHTML(b,D.budget[b],br);
    });
    h+='</div>'; el.innerHTML=h;
  }
}

function cardHTML(name,budget,r){
  var noBudget=!budget;
  var h='<div class="card fade-in"><div class="bu-name">'+name+'</div>';
  h+='<div class="row"><span class="lbl">年度费用(千元)</span><span class="val">'+(noBudget?'-':fmt(budget.ac))+'</span></div>';
  h+='<div class="row"><span class="lbl">年化HC(人)</span><span class="val">'+(noBudget?'-':fmt(budget.ah))+'</span></div>';
  h+='<div class="row"><span class="lbl">预测年度费用</span><span class="val predicted">'+fmt(r.totalCost)+'</span></div>';
  h+='<div class="row"><span class="lbl">预测年化HC</span><span class="val predicted">'+fmt(r.avgHC)+'</span></div>';
  if(!noBudget){
    var cp=parseFloat(r.cp),hp=parseFloat(r.hp);
    h+='<div class="progress-wrap"><div class="progress-label"><span>费用进度</span><span class="tag tag-'+pcolor(cp)+'">'+r.cp+'%</span></div>';
    h+='<div class="progress-bar"><div class="'+pclass(cp)+'" style="width:'+Math.min(cp,100)+'%"></div></div></div>';
    h+='<div class="progress-wrap"><div class="progress-label"><span>HC进度</span><span class="tag tag-'+pcolor(hp)+'">'+r.hp+'%</span></div>';
    h+='<div class="progress-bar"><div class="'+pclass(hp)+'" style="width:'+Math.min(hp,100)+'%"></div></div></div>';
  }
  h+='</div>';
  return h;
}

function renderDetail(bu,anchMonth,anchDate,rwData){
  var acData,ahData,rwCost,rwHC,chData,ccData;
  if(bu){
    var r=calcBU(bu,anchMonth,anchDate,rwData);
    acData=r.ac;ahData=r.ah;rwCost=r.rw.cost;rwHC=r.rw.hc;chData=r.ch;ccData=r.cc;
  } else {
    var r=calcAll(anchMonth,anchDate,rwData);
    acData=r.tac;ahData=r.tah;rwHC=r.rwHC;rwCost={};
    D.buList.forEach(function(bu){var c=(rwData[bu]||{}).cost||{};Object.keys(c).forEach(function(m){rwCost[m]=(rwCost[m]||0)+c[m]})});
    chData={};ccData={};
    for(var m=1;m<=12;m++){
      var ci=0,co=0,cii=0,coi=0;
      D.buList.forEach(function(b){ci+=gv(gv(D.ch[b],m),'intern');co+=gv(gv(D.ch[b],m),'cont');cii+=gv(gv(D.cc[b],m),'intern');coi+=gv(gv(D.cc[b],m),'cont')});
      chData[m]={intern:ci,cont:co};ccData[m]={intern:cii,cont:coi};
    }
  }
  // Actual
  var h='<div class="tbl-wrap"><table><thead><tr><th>指标</th>';
  for(var m=1;m<=12;m++) h+='<th>'+MNAME[m-1]+'</th>';
  h+='</tr></thead><tbody>';
  h+='<tr><td>费用(千元)</td>';
  for(var m=1;m<=12;m++){var dim=m>=anchMonth;var v=gv(acData,m);h+='<td class="'+(dim?'dim':'')+' num">'+(dim?'-':fmt(v))+'</td>';}
  h+='</tr><tr><td>人数</td>';
  for(var m=1;m<=12;m++){var dim=m>=anchMonth;var v=gv(ahData,m);h+='<td class="'+(dim?'dim':'')+' num">'+(dim?'-':fmt(v))+'</td>';}
  h+='</tr></tbody></table></div>';
  if(anchMonth<=1) h+='<div class="empty">锚点月为1月，无实际发生数据</div>';
  document.getElementById('tab-actual').innerHTML=h;
  // Renewal
  h='<div class="tbl-wrap"><table><thead><tr><th>指标</th>';
  for(var m=1;m<=12;m++) h+='<th>'+MNAME[m-1]+'</th>';
  h+='</tr></thead><tbody>';
  h+='<tr><td>费用(千元)</td>';
  for(var m=1;m<=12;m++){var dim=m>=anchMonth;var v=gv(rwCost,m);h+='<td class="'+(dim?'':'dim')+' num">'+(dim?fmt(v):'-')+'</td>';}
  h+='</tr><tr><td>人数</td>';
  for(var m=1;m<=12;m++){var dim=m>=anchMonth;var v=m>=anchMonth?rwHC:0;h+='<td class="'+(dim?'':'dim')+' num">'+(dim?fmt(v):'-')+'</td>';}
  h+='</tr></tbody></table></div>';
  document.getElementById('tab-renewal').innerHTML=h;
  // Invest
  h='<div class="tbl-wrap"><table><thead><tr><th>类型</th>';
  for(var m=1;m<=12;m++) h+='<th>'+MNAME[m-1]+'</th>';
  h+='</tr></thead><tbody>';
  var labels=['实习生转外包','继续投入'],keys=['intern','cont'];
  for(var i=0;i<2;i++){
    var lbl=labels[i],k=keys[i];
    h+='<tr><td>'+lbl+'(人)</td>';
    for(var m=1;m<=12;m++) h+='<td class="num">'+fmt(gv(gv(chData,m),k))+'</td>';
    h+='</tr><tr><td>'+lbl+'(千元)</td>';
    for(var m=1;m<=12;m++) h+='<td class="num">'+fmt(gv(gv(ccData,m),k))+'</td>';
    h+='</tr>';
  }
  h+='</tbody></table></div>';
  document.getElementById('tab-invest').innerHTML=h;
}

function renderPools(bu,anchMonth,anchDate,rwData){
  var sec=document.getElementById('poolSection'),el=document.getElementById('poolContent');
  if(!bu){sec.style.display='none';return}
  sec.style.display='block';
  var pools=D.pl[bu]||[];
  if(!pools.length){el.innerHTML='<div class="empty">该BU无资源池数据</div>';return}
  var rwAll=rwData.byPool||{};
  // 表头：展开图标 | 资源池代码 | 资源池名称 | 预算年费 | 预算HC
  //       | 实际发生年费 | 实际发生HC | 预测续签年费 | 预测续签HC
  var h='<div class="tbl-wrap"><table class="pool-table"><thead><tr>';
  h+='<th></th>';
  h+='<th>资源池代码</th>';
  h+='<th>资源池名称</th>';
  h+='<th>参照预算<br>(千元)</th>';
  h+='<th>预算HC<br>(人)</th>';
  h+='<th>预测年度<br>费用(千元)</th>';
  h+='<th style="color:var(--green)">实际发生<br>年费(千元)</th>';
  h+='<th style="color:var(--accent2)">预测续签<br>年费(千元)</th>';
  h+='</tr></thead><tbody>';
  pools.forEach(function(p,idx){
    var pc=p.code;
    var pac=D.acPl[pc]||{}, pah=D.ahPl[pc]||{};
    var prw=rwAll[pc]||{hc:0,cost:{}};
    var acCost=0;
    for(var m=1;m<anchMonth;m++) acCost+=gv(pac,m);
    var renCost=0;
    for(var m=anchMonth;m<=12;m++) renCost+=gv(prw.cost,m);
    var totalCost=acCost+renCost;
    var hasData=totalCost>0;
    h+='<tr class="pool-row" onclick="togglePool('+idx+')">';
    h+='<td style="color:var(--text3);width:30px">'+(hasData?'\u25B6':'-')+'</td>';
    h+='<td class="num" style="font-size:11px">'+pc+'</td>';
    h+='<td style="text-align:left;max-width:280px;overflow:hidden;text-overflow:ellipsis">'+p.name+'</td>';
    h+='<td class="num" style="color:var(--text3)">'+fmt(p.org_budget)+'</td>';
    h+='<td class="num">'+p.ahb.toFixed(1)+'</td>';
    h+='<td class="num">'+fmt(totalCost)+'</td>';
    h+='<td class="num" style="color:var(--green)">'+fmt(acCost)+'</td>';
    h+='<td class="num" style="color:var(--accent2)">'+fmt(renCost)+'</td></tr>';
    h+='<tr class="pool-detail" id="pd-'+idx+'"><td colspan="8"><div class="pool-detail-inner">'+poolDetailHTML(pc,pac,pah,prw,anchMonth)+'</div></td></tr>';
  });
  h+='</tbody></table></div>';
  el.innerHTML=h;
}

function poolDetailHTML(pc,pac,pah,prw,anchMonth){
  var h='<div style="display:flex;gap:20px;flex-wrap:wrap">';
  // 实际发生
  h+='<div class="sub-table"><div style="font-size:12px;color:var(--green);margin-bottom:4px;font-weight:600">&#9679; 实际发生</div><div class="tbl-wrap"><table><thead><tr><th>指标</th>';
  for(var m=1;m<=12;m++) h+='<th>'+MNAME[m-1]+'</th>';
  h+='</tr></thead><tbody><tr><td>费用</td>';
  for(var m=1;m<=12;m++){var dim=m>=anchMonth;var v=gv(pac,m);h+='<td class="'+(dim?'dim':'')+' num">'+(dim?'-':fmt(v))+'</td>';}
  h+='</tr><tr><td>人数</td>';
  for(var m=1;m<=12;m++){var dim=m>=anchMonth;var v=gv(pah,m);h+='<td class="'+(dim?'dim':'')+' num">'+(dim?'-':fmt(v))+'</td>';}
  h+='</tr></tbody></table></div></div>';
  // 预测续签
  h+='<div class="sub-table"><div style="font-size:12px;color:var(--accent2);margin-bottom:4px;font-weight:600">&#9679; 预测续签</div><div class="tbl-wrap"><table><thead><tr><th>指标</th>';
  for(var m=1;m<=12;m++) h+='<th>'+MNAME[m-1]+'</th>';
  h+='</tr></thead><tbody><tr><td>费用</td>';
  for(var m=1;m<=12;m++){var dim=m<anchMonth;var v=gv(prw.cost,m);h+='<td class="'+(dim?'dim':'')+' num">'+(dim?'-':fmt(v))+'</td>';}
  h+='</tr><tr><td>人数</td>';
  for(var m=1;m<=12;m++){var dim=m<anchMonth;var v=m>=anchMonth?prw.hc:0;h+='<td class="'+(dim?'dim':'')+' num">'+(dim?'-':fmt(v))+'</td>';}
  h+='</tr></tbody></table></div></div></div>';
  return h;
}

function togglePool(idx){var el=document.getElementById('pd-'+idx);if(el)el.classList.toggle('open')}
function switchTab(name){
  document.querySelectorAll('.tab').forEach(function(t){t.classList.remove('active')});
  document.querySelector('.tab[data-tab="'+name+'"]').classList.add('active');
  document.querySelectorAll('.tab-content').forEach(function(t){t.classList.remove('active')});
  document.getElementById('tab-'+name).classList.add('active');
}
function render(){
  var bu=document.getElementById('selBU').value;
  var anchRaw=document.getElementById('inpAnchor').value.trim();
  if(!/^\d{8}$/.test(anchRaw)){
    alert('请输入有效的锚点时间（格式：YYYYMMDD，如 20260410）');return;
  }
  var y=parseInt(anchRaw.substr(0,4));
  var m=parseInt(anchRaw.substr(4,2));
  var d=parseInt(anchRaw.substr(6,2));
  if(m<1||m>12||d<1||d>31){alert('日期无效，请检查');return;}
  var anchMonth=m;
  var anchDate=y+'-'+(m<10?'0':'')+m+'-'+(d<10?'0':'')+d;
  // 将锚点日期存入 sessionStorage，供 Excel 导出使用
  sessionStorage.setItem('anchDate', anchDate);
  sessionStorage.setItem('anchMonth', anchMonth);
  // 用精确日期计算续签数据
  var rw=calcRenewal(anchDate);
  renderCards(bu,anchMonth,anchDate,rw.byBU);
  renderDetail(bu,anchMonth,anchDate,rw.byBU);
  renderPools(bu,anchMonth,anchDate,rw);
}

// 点击「导出Excel」：纯客户端生成，调用 SheetJS 直接下载
function exportExcel(){
  var anch=sessionStorage.getItem('anchDate');
  if(!anch){alert('请先点击「查询」按钮设置锚点时间后再导出Excel');return;}
  var anchM=parseInt(anch.substr(5,2));

  // ---- 重新计算续签（与 calcRenewal 完全一致：锚点当天在岗口径）----
  var rwByBU={}, rwByPool={};
  D.staff.forEach(function(s){
    if(s.ws<=anch && s.we>=anch){
      var bu=s.bu;
      if(!rwByBU[bu]){rwByBU[bu]={hc:0,cost:{}};}
      rwByBU[bu].hc++;
      var pool=s.pool;
      if(pool){
        if(!rwByPool[pool]){rwByPool[pool]={hc:0,cost:{}};}
        rwByPool[pool].hc++;
      }
    }
  });
  for(var mm=anchM;mm<=12;mm++){
    var mi=mm-1;
    D.staff.forEach(function(s){
      if(s.ws<=anch && s.we>=anch){
        var price=s.mp[mi];
        var cost=Math.ceil(price*s.rate/100/1.06)/1000;
        if(rwByBU[s.bu]){
          rwByBU[s.bu].cost[mm]=(rwByBU[s.bu].cost[mm]||0)+cost;
        }
        if(s.pool && rwByPool[s.pool]){
          rwByPool[s.pool].cost[mm]=(rwByPool[s.pool].cost[mm]||0)+cost;
        }
      }
    });
  }
  Object.keys(rwByBU).forEach(function(bu){
    Object.keys(rwByBU[bu].cost).forEach(function(m){
      rwByBU[bu].cost[m]=Math.ceil(rwByBU[bu].cost[m]);
    });
  });

  // ---- 构建 Excel 数据 ----
  var BU_ORDER=['CMB','CSC','CMD','CUC','CTC','AID','ESU','AIO','RIC','CSU'];
  var ws_data=[];
  var MONTHS=__MONTHS__;
  // 表头
  var hdr=['BU','科目','单位'].concat(MONTHS.map(function(m){return 202600+m;}));
  ws_data.push(hdr);
  // 数据行
    BU_ORDER.forEach(function(bu){
    var rw=rwByBU[bu]||{hc:0,cost:{}};
    var ch_total={},cc_total={};
    MONTHS.forEach(function(m){
      var chm=(D.ch[bu]||{})[m]||{intern:0,cont:0};
      var ccm=(D.cc[bu]||{})[m]||{intern:0,cont:0};
      ch_total[m]=Math.ceil((chm.intern||0)+(chm.cont||0));
      cc_total[m]=Math.ceil((ccm.intern||0)+(ccm.cont||0));
    });
    var pushRow=function(base,monthsArr){
      var r=base.slice();   // [BU, 科目, 单位] 副本
      monthsArr.forEach(function(v){r.push(v);});
      ws_data.push(r);
    };
    pushRow([bu,'实际发生人数','人'],
      MONTHS.map(function(m){return m<anchM?(D.ah[bu]||{})[m]||0:null;}));
    pushRow([bu,'实际发生费用','￥K'],
      MONTHS.map(function(m){return m<anchM?(D.ac[bu]||{})[m]||0:null;}));
    pushRow([bu,'预测续签人数','人'],
      MONTHS.map(function(m){return m>=anchM?rw.hc:null;}));
    pushRow([bu,'预测续签费用','￥K'],
      MONTHS.map(function(m){return m>=anchM?(rw.cost[m]||0):null;}));
    pushRow([bu,'继续投入人数','人'],
      MONTHS.map(function(m){var v=ch_total[m];return v||null;}));
    pushRow([bu,'继续投入费用','￥K'],
      MONTHS.map(function(m){var v=cc_total[m];return v||null;}));
  });

  // ---- SheetJS 生成 ----
  // SheetJS 的 aoa_to_sheet 按列对齐：每行元素个数必须一致
  // ws_data 当前结构：每行 [BU, 科目, 单位, m1, m2, ...] ✓
  var ws=XLSX.utils.aoa_to_sheet(ws_data);

  // 设置列宽（A/B/C 列固定宽度，月份列等宽）
  ws['!cols']=[
    {wch:8},{wch:12},{wch:6}
  ].concat(MONTHS.map(function(){return {wch:10};}));

  var wb=XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb,ws,'第三方人员外包');
  XLSX.writeFile(wb,'滚动预测_'+anch.replace(/-/g,'')+'.xlsx');
}
render();
</script>
</body>
</html>'''

with open(OUT_FILE,'w',encoding='utf-8') as f:
    MONTHS_JS = json.dumps(list(range(1,13)))
    html = HTML_TMPL.replace('__JSON_PLACEHOLDER__', js) \
                    .replace('__ANCHORAGE__', ANCHORAGE_S) \
                    .replace('__MONTHS__', MONTHS_JS)
    f.write(html)
print(f"Done! Output: {OUT_FILE}")

# ============================================================
# 生成滚动预测 Excel（按锚点时间切割实际/预测）
# ============================================================
print("\nGenerating rolling forecast Excel...")
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment
import copy

def export_rolling_forecast(anch_date=None):
    if anch_date is None:
        anch_date = ANCHORAGE
    anch_month = anch_date.month
    anch_str = anch_date.strftime('%Y-%m-%d')
    out_name = f"{BASE_DIR}/滚动预测_{anch_date.strftime('%Y%m%d')}.xlsx"

    # 1. 计算预测续签（Python版，与JS calcRenewal一致）
    # 续签口径：锚点当天在岗人数，锚点当天在岗人员记录进行每月费用核算
    rw_bu = {}
    for s in staff_list:
        if s['ws'] <= anch_str <= s['we']:
            bu = s['bu']
            if bu not in rw_bu:
                rw_bu[bu] = {'hc': 0, 'cost': {m: 0.0 for m in range(1, 13)}}
            rw_bu[bu]['hc'] += 1
    for s in staff_list:
        if s['ws'] <= anch_str <= s['we']:
            bu = s['bu']
            for mm in range(anch_month, 13):
                price = s['mp'][mm - 1]
                cost = math.ceil(price * s['rate'] / 100 / 1.06) / 1000
                if bu in rw_bu:
                    rw_bu[bu]['cost'][mm] += cost
    for bu in rw_bu:
        for m in range(1, 13):
            rw_bu[bu]['cost'][m] = math.ceil(rw_bu[bu]['cost'][m])

    # 2. 加载模板，读取样式
    tmpl_path = f'{BASE_DIR}/滚动预测update.xlsx'
    wb_tmpl = load_workbook(tmpl_path)
    ws_tmpl = wb_tmpl.active

    def get_style(cell):
        return {
            'font': copy.copy(cell.font),
            'fill': copy.copy(cell.fill),
            'alignment': copy.copy(cell.alignment),
            'border': copy.copy(cell.border),
            'number_format': cell.number_format,
        }

    def apply_style(cell, style):
        cell.font = style['font']
        cell.fill = style['fill']
        cell.alignment = style['alignment']
        cell.border = style['border']
        cell.number_format = style['number_format']

    hdr_styles = [get_style(ws_tmpl.cell(1, c)) for c in range(1, 16)]
    dat_styles = [get_style(ws_tmpl.cell(2, c)) for c in range(1, 16)]

    # 3. 构建新工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = '第三方人员外包'

    for col_letter in ['A','B','C','D','E','F','G','H','I','J','K','L','M','N','O']:
        ws.column_dimensions[col_letter].width = ws_tmpl.column_dimensions[col_letter].width
    ws.row_dimensions[1].height = ws_tmpl.row_dimensions[1].height or 23.25

    # 表头
    headers = ['BU', '科目', '单位'] + [202600 + m for m in range(1, 13)]
    for c, val in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=val)
        apply_style(cell, hdr_styles[c - 1])

    # 数据行
    BU_ORDER = ['CMB','CSC','CMD','CUC','CTC','AID','ESU','AIO','RIC','CSU']
    MONTHS = list(range(1, 13))
    row_idx = 2

    for bu in BU_ORDER:
        ch_total = {m: ch_bu[bu][m].get('intern', 0) + ch_bu[bu][m].get('cont', 0) for m in MONTHS}
        cc_total = {m: cc_bu[bu][m].get('intern', 0) + cc_bu[bu][m].get('cont', 0) for m in MONTHS}
        rw = rw_bu.get(bu, {'hc': 0, 'cost': {m: 0 for m in MONTHS}})

        rows_def = [
            ('实际发生人数', '人',  {m: ah_bu[bu][m] if m < anch_month else None for m in MONTHS}),
            ('实际发生费用', '￥K', {m: ac_bu[bu][m] if m < anch_month else None for m in MONTHS}),
            ('预测续签人数', '人',  {m: rw['hc'] if m >= anch_month else None for m in MONTHS}),
            ('预测续签费用', '￥K', {m: rw['cost'].get(m, 0) if m >= anch_month else None for m in MONTHS}),
            ('继续投入人数', '人',  {m: ch_total[m] or None for m in MONTHS}),
            ('继续投入费用', '￥K', {m: cc_total[m] or None for m in MONTHS}),
        ]

        for subj, unit, mvals in rows_def:
            ws.row_dimensions[row_idx].height = 14
            ws.cell(row=row_idx, column=1, value=bu)
            ws.cell(row=row_idx, column=2, value=subj)
            ws.cell(row=row_idx, column=3, value=unit)
            for c in range(1, 4):
                apply_style(ws.cell(row=row_idx, column=c), dat_styles[c - 1])
            for mi, m in enumerate(MONTHS):
                v = mvals[m]
                cell = ws.cell(row=row_idx, column=4 + mi, value=v)
                apply_style(cell, dat_styles[3 + mi])
                cell.alignment = Alignment(horizontal='center')
            row_idx += 1

    wb.save(out_name)
    print(f"  Rolling forecast saved: {out_name}")
    return out_name

xl_out = export_rolling_forecast(anch_date=ANCHORAGE)
print(f"All done!")
