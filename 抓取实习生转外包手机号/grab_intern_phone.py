"""
实习生转外包手机号抓取工具
登录 work.asiainfo.com 联系人页面，根据 NT 账号逐一查询手机号并回写 Excel
"""

import re
import sys
import os
import argparse
from pathlib import Path
from typing import Optional

import openpyxl
from playwright.sync_api import sync_playwright, Page, Browser, Playwright


BASE_DIR = os.path.dirname(os.path.abspath(__file__))


class InternPhoneGrabber:
    """实习生转外包手机号抓取器"""

    CONTACT_URL = "https://work.asiainfo.com/zh/contact"

    SELECTORS = {
        "username_input": "input[name='username'], input[id='username'], input[type='text']",
        "password_input": "input[name='password'], input[id='password'], input[type='password']",
        "login_button": "button[type='submit'], input[type='submit'], button:has-text('登录'), "
                        "input[value*='登录'], a:has-text('登录'), button:has-text('Sign')",
    }

    def __init__(
        self,
        username: str,
        password: str,
        excel_path: str,
        headless: bool = False,
    ):
        self.username = username
        self.password = password
        self.excel_path = Path(excel_path)
        self.headless = headless

        if not self.excel_path.exists():
            raise FileNotFoundError(f"Excel 文件不存在: {self.excel_path}")

        self._playwright: Optional[Playwright] = None
        self._browser: Optional[Browser] = None
        self._page: Optional[Page] = None

    def start(self) -> None:
        self._playwright = sync_playwright().start()
        self._browser = self._playwright.chromium.launch(
            headless=self.headless,
        )
        context = self._browser.new_context(accept_downloads=True)
        self._page = context.new_page()
        print("[INFO] 浏览器已启动")

    def stop(self) -> None:
        if self._browser:
            self._browser.close()
        if self._playwright:
            self._playwright.stop()
        print("[INFO] 浏览器已关闭")

    def __enter__(self):
        self.start()
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        self.stop()

    # ===================== 登录 =====================

    def login(self) -> bool:
        print(f"[INFO] 正在访问联系人页面: {self.CONTACT_URL}")
        self._page.goto(self.CONTACT_URL, wait_until="networkidle")
        self._page.wait_for_timeout(3000)

        current_url = self._page.url

        # 检查是否被重定向到登录页
        if "login" in current_url.lower() or "auth" in current_url.lower() or \
           "signin" in current_url.lower() or "passport" in current_url.lower() or \
           "sso" in current_url.lower():
            return self._do_login()

        # 如果已经在联系人页面上但没有登录态，页面可能内嵌了登录表单
        if self._has_login_form():
            return self._do_login()

        print("[SUCCESS] 无需登录，已在联系人页面")
        return True

    def _has_login_form(self) -> bool:
        """检查当前页面是否有登录表单"""
        un_input = self._page.locator(self.SELECTORS["username_input"]).first
        return un_input.count() > 0

    def _do_login(self) -> bool:
        """执行登录操作"""
        print("[INFO] 正在执行登录...")

        # SSO 登录页可能有多步骤：先输入用户名 → 下一步 → 输入密码 → 登录
        username_input = self._page.locator(self.SELECTORS["username_input"]).first
        if username_input.count() == 0:
            print("[ERROR] 未找到用户名输入框")
            self._dump_page_text()
            return False

        username_input.click()
        username_input.fill(self.username)
        print(f"[INFO] 已输入用户名: {self.username}")
        self._page.wait_for_timeout(800)

        # 检查是否有"下一步"按钮（微软/ADFS SSO 常见）
        next_btn = self._page.locator(
            "input[value*='下一步'], button:has-text('下一步'), "
            "input[value*='Next'], button:has-text('Next'), "
            "#idSIButton9"
        ).first
        if next_btn.count() > 0:
            next_btn.click()
            print("[INFO] 已点击下一步")
            self._page.wait_for_timeout(3000)

            # 等待密码输入框出现
            password_input = self._page.locator(self.SELECTORS["password_input"]).first
            if password_input.count() > 0:
                password_input.click()
                password_input.fill(self.password)
                print("[INFO] 已输入密码")
                self._page.wait_for_timeout(800)

                # 点击登录/确认按钮
                submit_btn = self._page.locator(
                    "input[value*='登录'], button:has-text('登录'), "
                    "input[value*='Sign'], button:has-text('Sign'), "
                    "#idSIButton9, input[type='submit'], button[type='submit']"
                ).first
                if submit_btn.count() > 0:
                    submit_btn.click()
                    print("[INFO] 已点击登录按钮")
                    self._page.wait_for_timeout(5000)

            # 检查是否要求保持登录
            stay_btn = self._page.locator(
                "input[value*='是'], button:has-text('是'), "
                "input[value*='Yes'], button:has-text('Yes'), "
                "#idSIButton9"
            ).first
            if stay_btn.count() > 0:
                stay_btn.click()
                print("[INFO] 已点击保持登录")
                self._page.wait_for_timeout(5000)

        else:
            # 传统单页登录
            password_input = self._page.locator(self.SELECTORS["password_input"]).first
            if password_input.count() > 0:
                password_input.click()
                password_input.fill(self.password)
                print("[INFO] 已输入密码")
                self._page.wait_for_timeout(800)

            login_btn = self._page.locator(self.SELECTORS["login_button"]).first
            if login_btn.count() > 0:
                login_btn.click()
                print("[INFO] 已点击登录按钮")
                self._page.wait_for_timeout(5000)
            else:
                # 回车提交
                self._page.keyboard.press("Enter")
                print("[INFO] 已按回车提交")
                self._page.wait_for_timeout(5000)

        # 验证是否到达联系人页面
        current_url = self._page.url
        if "contact" in current_url.lower():
            print("[SUCCESS] 登录成功，已到达联系人页面")
            return True

        # 可能登录失败，再检查
        if self._has_login_form():
            print("[ERROR] 登录似乎未成功，页面仍显示登录表单")
            return False

        print("[INFO] 登录状态未确认，继续尝试...")
        return True

    def _dump_page_text(self):
        """打印页面文本用于调试"""
        try:
            text = self._page.locator("body").inner_text()
            print(f"[DEBUG] 页面文本(前500字): {text[:500]}")
        except Exception:
            pass

    # ===================== 页面有效性检查 =====================

    def _ensure_page_valid(self) -> bool:
        """检查当前页面是否有效，无效则重新导航到联系人页面"""
        try:
            _ = self._page.url
            return True
        except Exception:
            pass

        # 页面已关闭，重新创建
        print("[INFO] 页面已失效，重新创建...")
        try:
            new_page = self._browser.contexts[0].new_page()
            self._page = new_page
        except Exception:
            print("[ERROR] 无法重新创建页面")
            return False

        self._page.goto(self.CONTACT_URL, wait_until="networkidle")
        self._page.wait_for_timeout(3000)
        return True

    # ===================== 搜索联系人 =====================

    def search_contact(self, nt_account: str) -> Optional[str]:
        """
        在联系人页面搜索指定 NT 账号，返回手机号

        Args:
            nt_account: 亚信 NT 账号（如 jianglf3）

        Returns:
            Optional[str]: 手机号，未找到返回 None
        """
        print(f"[INFO] 正在搜索: {nt_account}")

        # 确保页面有效并导航到联系人页面
        if not self._ensure_page_valid():
            print(f"[ERROR] 页面无效 [{nt_account}]")
            return None

        self._page.goto(self.CONTACT_URL, wait_until="networkidle")
        self._page.wait_for_timeout(3000)

        # 定位搜索输入框
        search_input = self._find_search_input()
        if not search_input:
            print(f"[ERROR] 未找到搜索输入框 [{nt_account}]")
            self._dump_page_text()
            return None

        # 输入 NT 账号并搜索
        try:
            search_input.click()
            self._page.wait_for_timeout(300)
            search_input.fill(nt_account)
            print(f"[INFO] 已输入搜索关键词: {nt_account}")
            self._page.wait_for_timeout(500)

            # 优先按回车提交（避免按钮被 overlay 遮挡导致 click 失败）
            self._page.keyboard.press("Enter")
            print("[INFO] 已按回车提交搜索")
            self._page.wait_for_timeout(3000)
        except Exception as e:
            print(f"[ERROR] 搜索操作异常 [{nt_account}]: {e}")
            return None

        # 提取手机号
        try:
            phone = self._extract_phone_from_page()
        except Exception as e:
            print(f"[ERROR] 提取手机号异常 [{nt_account}]: {e}")
            phone = None

        if phone:
            print(f"[SUCCESS] 找到手机号: {phone} [{nt_account}]")
        else:
            print(f"[WARNING] 未找到手机号 [{nt_account}]")

        return phone

    def _find_search_input(self):
        """定位搜索输入框"""
        try:
            # 策略1：通过 placeholder 查找
            for placeholder in ["姓名/工号/NT", "姓名", "工号", "NT", "搜索", "Search"]:
                el = self._page.locator(f"input[placeholder*='{placeholder}']").first
                if el.count() > 0:
                    print(f"[DEBUG] 通过 placeholder 找到搜索输入框: {placeholder}")
                    return el

            # 策略2：第一个可见的 text/search input
            for el in self._page.locator("input[type='text'], input[type='search'], "
                                          "input:not([type])").all():
                if el.is_visible():
                    print("[DEBUG] 使用页面上第一个可见的文本输入框")
                    return el
        except Exception as e:
            print(f"[DEBUG] 查找搜索输入框异常: {e}")

        return None

    def _click_search_button(self) -> bool:
        """点击查询按钮（使用 force 绕过 overlay 遮挡）"""
        try:
            for text in ["查询", "搜索", "Search", "确定"]:
                btn = self._page.locator(
                    f"button:has-text('{text}'), a:has-text('{text}'), "
                    f"input[value*='{text}']"
                ).first
                if btn.count() > 0:
                    # 先用 JS dispatchEvent 绕过 overlay 拦截
                    btn.evaluate("el => el.click()")
                    print(f"[INFO] 已通过 JS 点击 '{text}' 按钮")
                    return True
        except Exception as e:
            print(f"[DEBUG] 点击搜索按钮异常: {e}")
        return False

    def _extract_phone_from_page(self) -> Optional[str]:
        """从当前页面搜索结果中提取手机号"""
        try:
            # 先尝试直接定位"手机/座机"所在的 DOM 元素
            phone_label = self._page.locator(
                "text=手机/座机, text=手机, text=座机"
            ).first
            if phone_label.count() > 0:
                parent_text = phone_label.locator("..").inner_text()
                patterns = [
                    r'手机[/\s]*(?:座机|电话)[：:]\s*(\d{11})',
                    r'手机[：:]\s*(\d{11})',
                    r'(?:座机|电话)[：:]\s*(\d{11})',
                ]
                for pat in patterns:
                    match = re.search(pat, parent_text)
                    if match and match.group(1).startswith("1"):
                        return match.group(1)
        except Exception:
            pass

        # 回退：遍历所有 body 元素获取文本（页面可能有多个 body，如 #app 内嵌）
        page_text = ""
        try:
            bodies = self._page.locator("body").all()
            for body in bodies:
                try:
                    page_text += body.inner_text() + "\n"
                except Exception:
                    continue
        except Exception:
            pass

        if not page_text:
            print("[DEBUG] 页面文本为空")
            return None

        print(f"[DEBUG] 页面文本(前300字): {page_text[:300]}")

        patterns = [
            r'手机[/\s]*(?:座机|电话)[：:]\s*(\d{11})',
            r'手机[：:]\s*(\d{11})',
            r'(?:座机|电话)[：:]\s*(\d{11})',
            r'联系方式[：:]\s*(\d{11})',
        ]

        for pattern in patterns:
            match = re.search(pattern, page_text)
            if match and match.group(1).startswith("1"):
                return match.group(1)

        # 最后一招：在页面中找第一个以1开头的11位数字
        if "手机" in page_text or "座机" in page_text or "电话" in page_text:
            match = re.search(r'(?<!\d)(1[3-9]\d{9})(?!\d)', page_text)
            if match:
                return match.group(1)

        return None

    # ===================== Excel 处理 =====================

    def process(self) -> str:
        """
        主处理流程：读取 Excel → 逐条查询 → 回写结果

        Returns:
            str: 输出文件路径
        """
        # 加载 Excel
        print(f"[INFO] 正在读取 Excel: {self.excel_path}")
        wb = openpyxl.load_workbook(self.excel_path)
        ws = wb.active
        print(f"[INFO] 工作表: {ws.title}, 行数={ws.max_row}, 列数={ws.max_column}")

        # 查找列索引
        header_row = 1
        nt_col = None
        phone_col = None
        headers = {}

        for col in range(1, ws.max_column + 1):
            val = ws.cell(row=header_row, column=col).value
            if val is None:
                continue
            # 清理：合并换行符、去空格
            clean_val = str(val).replace("\n", "").replace("\r", "").strip()
            headers[col] = clean_val

            if any(kw in clean_val for kw in ["NT", "nt", "邮箱", "賬号", "账号", "亚信"]):
                nt_col = col
            if "手机" in clean_val:
                phone_col = col

        print(f"[INFO] 列头: {headers}")
        if nt_col is None:
            print("[ERROR] 未找到 '邮箱/NT账号' 列")
            print(f"[DEBUG] 可用列: {list(headers.values())}")
            wb.close()
            sys.exit(1)
        if phone_col is None:
            print("[ERROR] 未找到 '手机号' 列")
            print(f"[DEBUG] 可用列: {list(headers.values())}")
            wb.close()
            sys.exit(1)

        print(f"[INFO] NT账号列: 第{nt_col}列, 手机号列: 第{phone_col}列")

        # 启动浏览器并登录
        self.start()
        try:
            if not self.login():
                print("[ERROR] 登录失败，无法继续")
                return ""

            success_count = 0
            fail_count = 0
            total = 0
            output_path = self._get_output_path()

            for row in range(2, ws.max_row + 1):
                nt_account = ws.cell(row=row, column=nt_col).value
                if nt_account is None or str(nt_account).strip() == "":
                    continue

                nt_account = str(nt_account).strip()
                existing_phone = ws.cell(row=row, column=phone_col).value

                # 跳过已有手机号的行
                if existing_phone and str(existing_phone).strip() not in ("", "None"):
                    print(f"[SKIP] 第{row}行 {nt_account} 已有手机号: {existing_phone}")
                    continue

                total += 1
                print(f"\n--- 处理第{row}行 [{total}]: {nt_account} ---")

                try:
                    phone = self.search_contact(nt_account)
                except Exception as e:
                    print(f"[ERROR] 第{row}行查询异常: {e}")
                    phone = None
                    # 尝试恢复页面状态
                    try:
                        if not self._ensure_page_valid():
                            print("[ERROR] 页面恢复失败，跳过此行")
                    except Exception:
                        pass

                if phone:
                    ws.cell(row=row, column=phone_col).value = phone
                    success_count += 1
                else:
                    ws.cell(row=row, column=phone_col).value = "失败"
                    fail_count += 1

                # 每查一条就保存一次（防丢失）
                try:
                    wb.save(output_path)
                except Exception as e:
                    print(f"[ERROR] 保存 Excel 失败: {e}")

        finally:
            self.stop()

        # 最终保存
        output_path = self._get_output_path()
        wb.save(output_path)
        wb.close()

        print(f"\n{'='*60}")
        print(f"  处理完成")
        print(f"{'='*60}")
        print(f"  总处理数: {total}")
        print(f"  成功: {success_count}")
        print(f"  失败: {fail_count}")
        print(f"  输出文件: {output_path}")
        print(f"{'='*60}")

        return output_path

    def _get_output_path(self) -> str:
        """获取输出文件路径（与输入同目录，添加 _已填手机号 后缀）"""
        stem = self.excel_path.stem
        suffix = self.excel_path.suffix
        parent = self.excel_path.parent
        return str(parent / f"{stem}_已填手机号{suffix}")


def main():
    """主函数 - 交互式输入"""
    try:
        from config import USERNAME, PASSWORD, EXCEL_PATH
    except ImportError:
        USERNAME = ""
        PASSWORD = ""
        EXCEL_PATH = ""

    parser = argparse.ArgumentParser(
        description="实习生转外包手机号抓取工具 — 从 work.asiainfo.com 联系人页面批量查询手机号"
    )
    parser.add_argument("-u", "--username", help="登录用户名", default=USERNAME)
    parser.add_argument("-p", "--password", help="登录密码", default=PASSWORD)
    parser.add_argument("-e", "--excel", help="Excel 文件路径", default=EXCEL_PATH)
    parser.add_argument("--headless", action="store_true", help="隐藏浏览器（无头模式）")

    args = parser.parse_args()

    username = args.username or input("请输入用户名: ")
    password = args.password or input("请输入密码: ")
    excel_path = args.excel or input("请输入 Excel 文件路径: ").strip()

    if not excel_path:
        print("[ERROR] 必须指定 Excel 文件路径")
        sys.exit(1)

    # 去除可能的引号
    excel_path = excel_path.strip('"').strip("'")

    if not os.path.isabs(excel_path):
        excel_path = os.path.abspath(excel_path)

    if not os.path.exists(excel_path):
        print(f"[ERROR] Excel 文件不存在: {excel_path}")
        sys.exit(1)

    print(f"\n{'='*60}")
    print(f"  实习生转外包手机号抓取工具")
    print(f"{'='*60}")
    print(f"  用户名: {username}")
    print(f"  Excel: {excel_path}")
    print(f"  无头模式: {args.headless}")
    print(f"{'='*60}\n")

    grabber = InternPhoneGrabber(
        username=username,
        password=password,
        excel_path=excel_path,
        headless=args.headless,
    )
    grabber.process()


if __name__ == "__main__":
    main()
