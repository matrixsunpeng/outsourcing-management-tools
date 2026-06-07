"""
Excel 待办清单解析工具
从续签查询报表中提取符合条件的待办清单
从人员变更报表中提取离岗不续签清单
"""

import pandas as pd
from pathlib import Path
from typing import Tuple, Optional
from datetime import datetime, date


class ExcelParser:
    """Excel 数据解析和待办清单生成"""

    # 必需的列名 - 续签查询报表
    REQUIRED_COLS = {
        "renewal_app": "待续签申请单",
        "status": "单据状态",
        "leave_time": "离职时间",
        "vendor_name": "技术合作商名称",
        "work_location": "工作地点"
    }

    # 人员变更报表列名
    PERSONNEL_COLS = {
        "leave_date": "预计离岗时间",
        "status": "单据状态",
        "id_card": "身份证"  # 可能是其他列名，如 "身份证号"、"idCard" 等
    }

    @staticmethod
    def _fix_id_card_columns(df: pd.DataFrame) -> pd.DataFrame:
        """
        修复被Excel转成科学计数法的身份证号/长数字列。

        问题：pandas读取Excel时，18位身份证号会被当作浮点数，显示为 3.10E+17 等科学计数法。
        修复：检测包含浮点数的疑似身份证列 → 转回字符串并补齐精度。
        """
        # 可能包含身份证号的列名关键词
        id_card_keywords = ["身份证", "idcard", "证件号", "id_card", "ID", "证件"]

        for col in df.columns:
            col_str = str(col).lower()

            # 只处理名称匹配的列
            matched_keyword = any(kw in col_str for kw in id_card_keywords)
            if not matched_keyword:
                continue

            series = df[col]
            if not pd.api.types.is_numeric_dtype(series):
                continue  # 已经是文本，无需修复

            # 抽样检查：非空值是否都是大数字（>= 10^14，即至少15位）
            sample = series.dropna().head(20)
            if len(sample) == 0:
                continue

            all_large_numbers = True
            for v in sample:
                try:
                    val = float(v)
                    if val < 1e14 or val > 1e19:  # 身份证号范围：15~19位
                        all_large_numbers = False
                        break
                except (ValueError, TypeError):
                    all_large_numbers = False
                    break

            if not all_large_numbers:
                continue

            # 转换：float → 整数字符串
            converted = []
            fixed_count = 0
            for idx in df.index:
                raw_val = series.iloc[idx]
                if pd.isna(raw_val):
                    converted.append("")
                    continue

                try:
                    f_val = float(raw_val)
                    s_val = f"{f_val:.0f}"  # 去掉小数，转为整数字符串
                    if len(s_val) >= 15 and len(s_val) <= 19:
                        converted.append(s_val)
                        fixed_count += 1
                    else:
                        converted.append(s_val)
                        fixed_count += 1
                except (ValueError, TypeError):
                    converted.append(str(raw_val))

            df[col] = converted
            print(f"[INFO] [身份证修复] 列 '{col}': 已将 {fixed_count} 条记录从科学计数法还原为文本")

        return df

    @staticmethod
    def _save_with_idcard_as_text(df: pd.DataFrame, output_file: Path) -> None:
        """
        保存DataFrame到Excel，确保身份证号等长数字列以文本格式写入。

        使用 openpyxl 底层 API 对疑似身份证列设置 Text 格式，
        防止Excel再次将其显示为科学计数法。
        """
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows

        # 疑似身份证列名关键词
        id_card_keywords = ["身份证", "idcard", "证件号", "id_card", "ID", "证件"]
        text_cols = {col for col in df.columns if any(kw in str(col).lower() for kw in id_card_keywords)}

        wb = Workbook()
        ws = wb.active
        ws.title = "Sheet1"

        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True)):
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx + 1, column=c_idx, value=value)

        # 对身份证列设置文本格式（仅对数据行，跳过表头）
        header_row = list(df.columns)
        for col_idx, col_name in enumerate(header_row, start=1):
            if col_name in text_cols:
                from openpyxl.styles import numbers
                for row_idx in range(2, len(df) + 2):  # 第1行是表头
                    ws.cell(row=row_idx, column=col_idx).number_format = numbers.FORMAT_TEXT

        wb.save(output_file)

    @staticmethod
    def parse_application_form(file_path: str) -> set:
        """
        解析下载的"申请单"Excel，提取"合作申请单编号"列的所有值。

        Excel格式：第1行为报表标题，第2行为列名，数据从第3行开始。
        使用 skiprows=1 跳过后 header=0（第2行作为列名）。

        Args:
            file_path: 申请单 xlsx 文件路径

        Returns:
            set: 合作申请单编号集合（已去重、去空格、去空值）
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"申请单文件不存在: {file_path}")

        print(f"[INFO] 正在解析申请单: {file_path}")
        df = pd.read_excel(file_path, sheet_name=0, skiprows=1, dtype=str)
        df = df.replace(["nan", "NaN", "NAN", "", "None", "none"], "")

        print(f"[INFO] 申请单原始记录数: {len(df)}")
        print(f"[INFO] 申请单列名: {list(df.columns)}")

        # 查找"合作申请单编号"列
        app_no_col = None
        for col_name in ["合作申请单编号", "申请单编号", "合作申请单号"]:
            if col_name in df.columns:
                app_no_col = col_name
                break

        if not app_no_col:
            raise ValueError(f"申请单中未找到'合作申请单编号'列。可用列: {list(df.columns)}")

        # 提取并清理
        app_nos = set(
            str(v).strip()
            for v in df[app_no_col].dropna().values
            if str(v).strip()
        )
        print(f"[INFO] 申请单中有效合作申请单编号数: {len(app_nos)}")

        return app_nos

    @staticmethod
    def load_and_filter(file_path: str,
                        valid_app_nos: Optional[set] = None) -> pd.DataFrame:
        """
        加载并筛选 Excel 文件

        Args:
            file_path: xlsx 文件路径
            valid_app_nos: 有效申请单号集合（来自申请单），用于额外过滤

        Returns:
            pd.DataFrame: 筛选后的数据框

        Raises:
            FileNotFoundError: 文件不存在
            ValueError: 缺少必需列
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        print(f"[INFO] 正在加载文件: {file_path}")
        # ★ 关键：dtype=str 强制所有列读为文本，防止18位身份证号被转成float丢失精度
        df = pd.read_excel(file_path, sheet_name=0, skiprows=1, dtype=str)

        # 清理全空字符串的单元格（pandas读入时空值可能变成 "nan" 或 ""）
        df = df.replace(["nan", "NaN", "NAN", "", "None", "none"], "")

        print(f"[INFO] 原始记录数: {len(df)}")
        print(f"[INFO] 列名: {list(df.columns)}")

        # 验证必需列
        missing_cols = [col for col in ExcelParser.REQUIRED_COLS.values() if col not in df.columns]
        if missing_cols:
            raise ValueError(f"缺少必需列: {missing_cols}\n可用列: {list(df.columns)}")

        # 条件1: "待续签申请单" 不为空
        cond1 = df[ExcelParser.REQUIRED_COLS["renewal_app"]].notna() & \
                (df[ExcelParser.REQUIRED_COLS["renewal_app"]] != "")

        # 条件2: "单据状态" == "审批流程结束"
        cond2 = df[ExcelParser.REQUIRED_COLS["status"]] == "审批流程结束"

        # 条件3: "离职时间" 为空
        cond3 = df[ExcelParser.REQUIRED_COLS["leave_time"]].isna() | \
                (df[ExcelParser.REQUIRED_COLS["leave_time"]] == "")

        # 综合条件筛选
        filtered_df = df[cond1 & cond2 & cond3].copy()
        print(f"[INFO] 基础筛选后记录数: {len(filtered_df)}")

        # 条件4 (可选): 只保留"待续签申请单"在申请单有效编号集合内的记录
        if valid_app_nos is not None and len(valid_app_nos) > 0 and len(filtered_df) > 0:
            before = len(filtered_df)
            renewal_col = ExcelParser.REQUIRED_COLS["renewal_app"]
            cond4 = filtered_df[renewal_col].apply(
                lambda x: str(x).strip() in valid_app_nos
            )
            filtered_df = filtered_df[cond4].copy()
            print(f"[INFO] 申请单匹配筛选: {before} → {len(filtered_df)} 条")

        return filtered_df

    @staticmethod
    def generate_todo_list(filtered_df: pd.DataFrame) -> pd.DataFrame:
        """
        生成待办清单：保留所有符合条件的记录，不按待续签申请单去重

        说明：
        - 下载的续签查询报表中，可能存在"待续签申请单"相同但代表不同明细行的多条记录
        - 只要记录本身符合筛选条件，就必须全部保留，不能分组后只留一条

        Args:
            filtered_df: 筛选后的数据框

        Returns:
            pd.DataFrame: 待办清单（包含反馈列，保留所有原始字段和原始行粒度）
        """
        # 保留所有符合条件的原始记录，维持原始顺序
        todo_df = filtered_df.copy().reset_index(drop=True)

        # 增加"反馈"列（初始为空）
        todo_df["反馈"] = ""

        print(f"[INFO] 生成待办清单记录数: {len(todo_df)}")
        print(f"[INFO] 待办清单列数: {len(todo_df.columns)}")
        print(f"[INFO] 清单列名: {list(todo_df.columns)}")

        return todo_df

    @staticmethod
    def save_todo_list(todo_df: pd.DataFrame, output_dir: str = ".") -> str:
        """
        保存待办清单到 Excel 文件

        Args:
            todo_df: 待办清单数据框
            output_dir: 输出目录

        Returns:
            str: 输出文件路径
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"待办订单清单_{timestamp}.xlsx"

        # ★ 用自定义保存方法，身份证号列写为文本格式
        ExcelParser._save_with_idcard_as_text(todo_df, output_file)
        print(f"[SUCCESS] 待办清单已保存: {output_file}")

        return str(output_file)

    @staticmethod
    def update_status(todo_df: pd.DataFrame, idx: int, status: str) -> None:
        """
        更新指定行的反馈状态

        Args:
            todo_df: 待办清单数据框
            idx: 行索引
            status: 状态文本
        """
        # 兼容旧列名"处理状态"
        col = "反馈" if "反馈" in todo_df.columns else "处理状态"
        todo_df[col] = todo_df[col].astype(str)
        todo_df.loc[idx, col] = status

    @staticmethod
    def save_result(todo_df: pd.DataFrame, output_dir: str = ".") -> str:
        """
        保存处理完成的待办清单

        Args:
            todo_df: 待办清单数据框（含处理状态）
            output_dir: 输出目录

        Returns:
            str: 输出文件路径
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"待办订单清单_已处理_{timestamp}.xlsx"

        # ★ 用自定义保存方法，身份证号列写为文本格式
        ExcelParser._save_with_idcard_as_text(todo_df, output_file)
        print(f"[SUCCESS] 处理结果已保存: {output_file}")

        return str(output_file)

    # ===================== 人员变更清单处理 =====================

    @staticmethod
    def load_personnel_change(file_path: str) -> pd.DataFrame:
        """
        加载人员变更Excel文件

        Args:
            file_path: xlsx 文件路径

        Returns:
            pd.DataFrame: 加载后的数据框

        Raises:
            FileNotFoundError: 文件不存在
        """
        file_path = Path(file_path)
        if not file_path.exists():
            raise FileNotFoundError(f"文件不存在: {file_path}")

        print(f"[INFO] 正在加载人员变更文件: {file_path}")
        # 跳过第一行（标题行），第二行作为列名
        # ★ dtype=str 强制文本模式，防止身份证号被转float丢失精度
        df = pd.read_excel(file_path, sheet_name=0, skiprows=1, dtype=str)
        df = df.replace(["nan", "NaN", "NAN", "", "None", "none"], "")

        print(f"[INFO] 原始人员变更记录数: {len(df)}")
        print(f"[INFO] 列名: {list(df.columns)}")

        return df

    @staticmethod
    def filter_personnel_not_renewing(df: pd.DataFrame, audit_date: Optional[str] = None) -> pd.DataFrame:
        """
        筛选离岗不续签的人员清单

        条件：
        1. 预计离岗时间 <= 稽核时间（续签查询结束时间），且 >= 稽核时间前20天
        2. 单据状态 = "审批流程中" 或 "审批流程结束"
        3. 去重：相同人员多条记录保留预计离岗时间最迟的

        Args:
            df: 人员变更数据框
            audit_date: 稽核时间（格式：YYYY-MM-DD或YYYY年MM月DD日），默认为当前日期

        Returns:
            pd.DataFrame: 筛选后的离岗不续签清单
        """
        from datetime import timedelta
        # 解析稽核时间
        if audit_date is None:
            audit_date = datetime.now().strftime("%Y-%m-%d")
        else:
            # 支持多种日期格式
            if "年" in audit_date:
                audit_date = audit_date.replace("年", "-").replace("月", "-").replace("日", "")

        try:
            audit_date_obj = pd.to_datetime(audit_date)
        except Exception as e:
            raise ValueError(f"稽核时间格式错误: {audit_date}，错误: {e}")

        # 计算20天前的日期（下限）
        date_lower = audit_date_obj - pd.Timedelta(days=20)

        print(f"[INFO] ========== 离岗不续签清单筛选 ==========")
        print(f"[INFO] 原始记录总数: {len(df)}")
        print(f"[INFO] 稽核时间（上限）: {audit_date_obj.date()}")
        print(f"[INFO] 预计离岗时间下限（前20天）: {date_lower.date()}")

        # 寻找预计离岗时间列（可能的列名）
        leave_date_col = None
        for col_name in ["预计离岗时间", "离岗时间", "预计离岗日期", "离岗日期"]:
            if col_name in df.columns:
                leave_date_col = col_name
                break

        if not leave_date_col:
            raise ValueError(f"未找到预计离岗时间列。可用列: {list(df.columns)}")

        # 寻找单据状态列
        status_col = None
        for col_name in ["单据状态", "状态"]:
            if col_name in df.columns:
                status_col = col_name
                break

        if not status_col:
            raise ValueError(f"未找到单据状态列。可用列: {list(df.columns)}")

        # 条件1: 预计离岗时间在范围内（不晚于稽核时间，不早于稽核时间前20天）
        df[leave_date_col] = pd.to_datetime(df[leave_date_col], errors='coerce')
        cond1 = (df[leave_date_col] >= date_lower) & (df[leave_date_col] <= audit_date_obj)

        print(f"[INFO] 时间范围筛选（{date_lower.date()} ~ {audit_date_obj.date()}）后: {cond1.sum()} 条")

        # 条件2: 单据状态为"审批流程中"或"审批流程结束"
        cond2 = df[status_col].isin(["审批流程中", "审批流程结束"])

        print(f"[INFO] 单据状态筛选后: {(cond1 & cond2).sum()} 条")

        filtered_df = df[cond1 & cond2].copy()
        print(f"[INFO] 符合条件的人员变更记录数: {len(filtered_df)}")

        # 条件3: 去重 - 相同人员保留预计离岗时间最迟的
        # 寻找身份证列
        id_card_col = None
        for col_name in ["身份证", "身份证号", "idCard", "证件号", "ID"]:
            if col_name in filtered_df.columns:
                id_card_col = col_name
                break

        if id_card_col:
            # 按身份证分组，保留预计离岗时间最迟的
            filtered_df = filtered_df.sort_values(leave_date_col, ascending=False)
            filtered_df = filtered_df.drop_duplicates(subset=[id_card_col], keep='first')
            print(f"[INFO] 去重后离岗不续签人员记录数: {len(filtered_df)}")
        else:
            print(f"[WARNING] 未找到身份证列，跳过去重。可用列: {list(filtered_df.columns)}")

        return filtered_df

    @staticmethod
    def save_personnel_not_renewing_list(not_renewing_df: pd.DataFrame, output_dir: str = ".") -> str:
        """
        保存离岗不续签清单到 Excel 文件

        Args:
            not_renewing_df: 离岗不续签清单数据框
            output_dir: 输出目录

        Returns:
            str: 输出文件路径
        """
        output_dir = Path(output_dir)
        output_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = output_dir / f"离岗不续签清单_{timestamp}.xlsx"

        # ★ 用自定义保存方法，身份证号列写为文本格式
        ExcelParser._save_with_idcard_as_text(not_renewing_df, output_file)
        print(f"[SUCCESS] 离岗不续签清单已保存: {output_file}")

        return str(output_file)


def main():
    """测试脚本"""
    import sys

    if len(sys.argv) < 2:
        print("使用方法: python excel_parser.py <xlsx文件路径>")
        sys.exit(1)

    file_path = sys.argv[1]

    try:
        # 加载并筛选
        filtered_df = ExcelParser.load_and_filter(file_path)

        # 生成待办清单
        todo_df = ExcelParser.generate_todo_list(filtered_df)

        # 保存待办清单
        output_file = ExcelParser.save_todo_list(todo_df, output_dir=".")

        print(f"\n[INFO] 待办清单内容:\n{todo_df.to_string(index=False)}")

    except Exception as e:
        print(f"[ERROR] 处理失败: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
